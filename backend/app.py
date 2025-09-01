from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text
from flask_cors import CORS
from datetime import datetime, date, timedelta, time
import os
import json
import logging
import io
from dotenv import load_dotenv

# Business Day Helper Functions
def get_business_day_from_datetime(dt):
    """
    Convert a datetime to the corresponding business day.
    Business day runs from 8:30 AM to 8:30 AM next day.
    Returns the date of the business day.
    """
    business_day_start_time = time(8, 30)  # 8:30 AM
    
    if dt.time() >= business_day_start_time:
        # If time is 8:30 AM or later, it belongs to current date's business day
        return dt.date()
    else:
        # If time is before 8:30 AM, it belongs to previous date's business day
        return (dt - timedelta(days=1)).date()

def get_current_business_day():
    """Get the current business day based on current time."""
    return get_business_day_from_datetime(datetime.now())

def is_night_shift_time():
    """Check if current time is during night shift (8:30 PM to 8:30 AM)."""
    current_time = datetime.now().time()
    night_start = time(20, 30)  # 8:30 PM
    day_start = time(8, 30)     # 8:30 AM
    
    # Night shift spans midnight, so check if current time is after 8:30 PM or before 8:30 AM
    return current_time >= night_start or current_time < day_start

def get_business_day_range(business_date):
    """
    Get the datetime range for a business day.
    Returns (start_datetime, end_datetime) for the business day.
    """
    start_datetime = datetime.combine(business_date, time(8, 30))
    end_datetime = datetime.combine(business_date + timedelta(days=1), time(8, 30))
    return start_datetime, end_datetime

# Import OpenAI - handle both old and new versions
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("OpenAI not available")

# LangChain imports for SQL Agent
try:
    # Temporarily disable langchain imports due to compatibility issues
    # from langchain_openai import ChatOpenAI
    # from langchain_community.utilities import SQLDatabase
    # from langchain_community.agent_toolkits import SQLDatabaseToolkit
    # from langchain_community.agent_toolkits.sql.base import create_sql_agent
    # from langchain.agents.agent_types import AgentType
    # from langchain.schema import SystemMessage
    LANGCHAIN_AVAILABLE = False
    print("LangChain modules temporarily disabled due to compatibility issues")
except ImportError as e:
    LANGCHAIN_AVAILABLE = False
    print(f"LangChain not available: {e}")

# Google Sheets integration imports - DISABLED due to quota issues
# try:
#     from google_drive_integration import GoogleDriveDBSync, GoogleSheetsDataProvider
#     from google_sheets_ai_chat import GoogleSheetsAIChat
#     GOOGLE_SHEETS_AVAILABLE = True
#     print("Google Sheets integration imported successfully")
# except ImportError as e:
GOOGLE_SHEETS_AVAILABLE = False
print("Google Sheets integration disabled to avoid quota errors")

# Load environment variables from .env file
load_dotenv()

# Setup logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///petrol_station.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Ensure the instance folder exists
try:
    os.makedirs(app.instance_path)
except OSError:
    pass

db = SQLAlchemy(app)

# Configure CORS to allow all origins and methods
CORS(app, supports_credentials=True, origins="*")

# Set OpenAI API key and configuration
openai_api_key = os.getenv('OPENAI_API_KEY')
if not openai_api_key:
    logger.error("OpenAI API key not found in environment variables")
    # For development only - set your key in .env file
    logger.info("Using environment variable or .env file for API key")

# Enable debug mode for testing - will use mock responses instead of calling the API
USE_MOCK_RESPONSES = os.getenv('USE_MOCK_RESPONSES', 'False').lower() == 'true'
logger.info(f"Mock responses {'enabled' if USE_MOCK_RESPONSES else 'disabled'}")

# Configure OpenAI client - handle both new and old API versions
try:
    # For all versions, set legacy API key
    openai.api_key = openai_api_key
    
    # Only try to use the API if not using mock responses
    if not USE_MOCK_RESPONSES:
        # Try the newer OpenAI package version first (v1.0.0+)
        try:
            logger.info("Trying to initialize newer OpenAI client...")
            client = openai.OpenAI(api_key=openai_api_key)
            
            # Test connection with newer client
            test_response = client.chat.completions.create(
                model="gpt-4",
                messages=[{"role": "user", "content": "Test"}],
                max_completion_tokens=5
            )
            logger.info("Successfully initialized newer OpenAI client")
        except (ImportError, AttributeError):
            # Try the older OpenAI package version (v0.x)
            logger.info("Newer client failed, trying legacy OpenAI client...")
            
            # Test connection with legacy client
            test_response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": "Test"}],
                max_completion_tokens=5
            )
            logger.info("Successfully initialized legacy OpenAI client")
            
        logger.info(f"OpenAI API connection verified")
        logger.info(f"OpenAI client initialized with key starting with: {openai_api_key[:5]}...")
    else:
        logger.info("Using mock responses for OpenAI - no API calls will be made")
except Exception as e:
    logger.error(f"Failed to initialize OpenAI client or verify connection: {str(e)}")
    logger.info("Using mock responses for OpenAI as fallback")
    # Set USE_MOCK_RESPONSES to True as fallback
    USE_MOCK_RESPONSES = True
    USE_MOCK_RESPONSES = True

# Database Models
class DailyConsolidation(db.Model):
    __tablename__ = 'daily_consolidation'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    shift = db.Column(db.String(10), nullable=False)  # Day/Night
    manager = db.Column(db.String(100), nullable=False)
    
    # Fuel Sales
    ms_rate = db.Column(db.Float, nullable=False, default=0.0)
    ms_quantity = db.Column(db.Float, nullable=False, default=0.0)
    ms_amount = db.Column(db.Float, nullable=False, default=0.0)
    
    hsd_rate = db.Column(db.Float, nullable=False, default=0.0)
    hsd_quantity = db.Column(db.Float, nullable=False, default=0.0)
    hsd_amount = db.Column(db.Float, nullable=False, default=0.0)
    
    power_rate = db.Column(db.Float, nullable=False, default=0.0)
    power_quantity = db.Column(db.Float, nullable=False, default=0.0)
    power_amount = db.Column(db.Float, nullable=False, default=0.0)
    
    # Tank Levels
    hsd1_tank = db.Column(db.Float, nullable=False, default=0.0)
    hsd2_tank = db.Column(db.Float, nullable=False, default=0.0)
    ms1_tank = db.Column(db.Float, nullable=False, default=0.0)
    ms2_tank = db.Column(db.Float, nullable=False, default=0.0)
    power1_tank = db.Column(db.Float, nullable=False, default=0.0)
    
    # HPCL Credit
    total_outstanding = db.Column(db.Float, nullable=False, default=0.0)
    hpcl_payment = db.Column(db.Float, nullable=False, default=0.0)  # New field for HPCL payments
    
    # Collections
    cash_collections = db.Column(db.Float, nullable=False, default=0.0)
    card_collections = db.Column(db.Float, nullable=False, default=0.0)
    paytm_collections = db.Column(db.Float, nullable=False, default=0.0)
    hp_transactions = db.Column(db.Float, nullable=False, default=0.0)
    
    # Manager Notes
    manager_notes = db.Column(db.Text)
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.isoformat(),
            'shift': self.shift,
            'manager': self.manager,
            'ms_rate': self.ms_rate,
            'ms_quantity': self.ms_quantity,
            'ms_amount': self.ms_amount,
            'hsd_rate': self.hsd_rate,
            'hsd_quantity': self.hsd_quantity,
            'hsd_amount': self.hsd_amount,
            'power_rate': self.power_rate,
            'power_quantity': self.power_quantity,
            'power_amount': self.power_amount,
            'hsd1_tank': self.hsd1_tank,
            'hsd2_tank': self.hsd2_tank,
            'ms1_tank': self.ms1_tank,
            'ms2_tank': self.ms2_tank,
            'power1_tank': self.power1_tank,
            'total_outstanding': self.total_outstanding,
            'hpcl_payment': self.hpcl_payment,
            'cash_collections': self.cash_collections,
            'card_collections': self.card_collections,
            'paytm_collections': self.paytm_collections,
            'hp_transactions': self.hp_transactions,
            'manager_notes': self.manager_notes,
            'created_at': self.created_at.isoformat()
        }

class ProcurementData(db.Model):
    __tablename__ = 'procurement_data'
    
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(100), nullable=False)
    invoice_date = db.Column(db.Date, nullable=False)
    fuel_type = db.Column(db.String(20), nullable=False)  # MS/HSD/POWER
    quantity = db.Column(db.Float, nullable=False)
    rate = db.Column(db.Float, nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    vehicle_number = db.Column(db.String(50), nullable=False)
    supplier = db.Column(db.String(100), default='HPCL')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'invoice_number': self.invoice_number,
            'invoice_date': self.invoice_date.isoformat(),
            'fuel_type': self.fuel_type,
            'quantity': self.quantity,
            'rate': self.rate,
            'total_amount': self.total_amount,
            'vehicle_number': self.vehicle_number,
            'supplier': self.supplier,
            'created_at': self.created_at.isoformat()
        }

class ChatHistory(db.Model):
    __tablename__ = 'chat_history'
    
    id = db.Column(db.Integer, primary_key=True)
    user_message = db.Column(db.Text, nullable=False)
    ai_response = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_message': self.user_message,
            'ai_response': self.ai_response,
            'timestamp': self.timestamp.isoformat()
        }

class TankReading(db.Model):
    __tablename__ = 'tank_readings'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    
    # Tank Levels
    hsd1_tank = db.Column(db.Float, nullable=False, default=0.0)
    hsd2_tank = db.Column(db.Float, nullable=False, default=0.0)
    ms1_tank = db.Column(db.Float, nullable=False, default=0.0)
    ms2_tank = db.Column(db.Float, nullable=False, default=0.0)
    power1_tank = db.Column(db.Float, nullable=False, default=0.0)
    
    # Notes
    notes = db.Column(db.Text)
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.isoformat(),
            'hsd1_tank': self.hsd1_tank,
            'hsd2_tank': self.hsd2_tank,
            'ms1_tank': self.ms1_tank,
            'ms2_tank': self.ms2_tank,
            'power1_tank': self.power1_tank,
            'notes': self.notes,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

class CustomerCredit(db.Model):
    __tablename__ = 'customer_credit'
    
    id = db.Column(db.Integer, primary_key=True)
    customer_name = db.Column(db.String(100), nullable=False)
    date = db.Column(db.Date, nullable=False)
    fuel_type = db.Column(db.String(20), nullable=False)  # MS/HSD/POWER
    quantity = db.Column(db.Float, nullable=False)
    rate = db.Column(db.Float, nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # 'sale' or 'payment'
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'customer_name': self.customer_name,
            'date': self.date.isoformat(),
            'fuel_type': self.fuel_type,
            'quantity': self.quantity,
            'rate': self.rate,
            'total_amount': self.total_amount,
            'transaction_type': self.transaction_type,
            'notes': self.notes,
            'created_at': self.created_at.isoformat()
        }

class HPCLPayments(db.Model):
    __tablename__ = 'hpcl_payments'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(50), default='Bank Transfer')  # Bank Transfer, Cash, Check, etc.
    reference_number = db.Column(db.String(100))  # Bank reference or check number
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.isoformat(),
            'amount': self.amount,
            'payment_method': self.payment_method,
            'reference_number': self.reference_number,
            'notes': self.notes,
            'created_at': self.created_at.isoformat()
        }

# Initialize database tables
with app.app_context():
    db.create_all()
    logger.info("Database tables created successfully")
    
    # Test database connection
    try:
        db.session.execute(text('SELECT 1'))  # Use text() here
        logger.info("Database connection successful")
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")

# Initialize Google Sheets Integration - COMPLETELY DISABLED
google_sync = None
google_ai_chat = None

# Google integration disabled due to quota issues
logger.info("Google Sheets integration not available")

# Helper Functions
def get_complete_business_data():
    """Get all business data for AI analysis"""
    daily_data = DailyConsolidation.query.all()
    procurement_data = ProcurementData.query.all()
    
    return {
        'daily_entries': [entry.to_dict() for entry in daily_data],
        'procurement_entries': [entry.to_dict() for entry in procurement_data],
        'total_daily_entries': len(daily_data),
        'total_procurement_entries': len(procurement_data)
    }

def get_business_data_summary():
    """Generates a detailed summary of business data for the AI prompt."""
    summary = {}
    today = date.today()

    # Daily entries summary
    daily_entries_count = db.session.query(DailyConsolidation).count()
    if daily_entries_count > 0:
        # Basic summary
        newest_date = db.session.query(func.max(DailyConsolidation.date)).scalar()
        oldest_date = db.session.query(func.min(DailyConsolidation.date)).scalar()
        
        summary['daily_data_summary'] = {
            'total_entries': daily_entries_count,
            'date_range': {
                'oldest': oldest_date.isoformat(),
                'newest': newest_date.isoformat()
            },
            'total_sales': db.session.query(func.sum(DailyConsolidation.ms_amount + DailyConsolidation.hsd_amount + DailyConsolidation.power_amount)).scalar(),
            'total_collections': db.session.query(func.sum(DailyConsolidation.cash_collections + DailyConsolidation.card_collections + DailyConsolidation.paytm_collections + DailyConsolidation.hp_transactions)).scalar(),
            'total_ms_quantity_sold': db.session.query(func.sum(DailyConsolidation.ms_quantity)).scalar(),
            'total_hsd_quantity_sold': db.session.query(func.sum(DailyConsolidation.hsd_quantity)).scalar(),
            'total_power_quantity_sold': db.session.query(func.sum(DailyConsolidation.power_quantity)).scalar(),
        }
        
        # Get latest tank levels
        latest_entry = DailyConsolidation.query.order_by(DailyConsolidation.date.desc()).first()
        if latest_entry:
            summary['current_tank_levels'] = {
                'hsd1_tank': latest_entry.hsd1_tank,
                'hsd2_tank': latest_entry.hsd2_tank,
                'ms1_tank': latest_entry.ms1_tank,
                'ms2_tank': latest_entry.ms2_tank,
                'power1_tank': latest_entry.power1_tank,
                'as_of_date': latest_entry.date.isoformat(),
            }
            
            # Add tank capacities
            summary['tank_capacities'] = {
                'hsd1_tank': 16000,  # KL
                'hsd2_tank': 22000,  # KL
                'ms1_tank': 9000,    # KL
                'ms2_tank': 9000,    # KL
                'power1_tank': 9000  # KL
            }
            
            # Add current fuel rates from the latest daily consolidation entry
            try:
                # Use rates directly from the latest daily entry
                summary['current_fuel_rates'] = {
                    'ms_rate': latest_entry.ms_rate,
                    'hsd_rate': latest_entry.hsd_rate,
                    'power_rate': latest_entry.power_rate,
                    'as_of_date': latest_entry.date.isoformat(),
                    'source': 'daily_consolidation'
                }
                logger.info(f"Using fuel rates from daily consolidation: MS={latest_entry.ms_rate}, HSD={latest_entry.hsd_rate}, POWER={latest_entry.power_rate}")
                
                # Also add procurement rates for reference and comparison
                try:
                    ms_rate = db.session.query(ProcurementData.rate).filter(ProcurementData.fuel_type == 'MS').order_by(ProcurementData.invoice_date.desc()).first()
                    hsd_rate = db.session.query(ProcurementData.rate).filter(ProcurementData.fuel_type == 'HSD').order_by(ProcurementData.invoice_date.desc()).first()
                    power_rate = db.session.query(ProcurementData.rate).filter(ProcurementData.fuel_type == 'POWER').order_by(ProcurementData.invoice_date.desc()).first()
                    
                    procurement_rates = {}
                    if ms_rate and ms_rate[0]:
                        procurement_rates['ms_rate'] = float(ms_rate[0])
                    if hsd_rate and hsd_rate[0]:
                        procurement_rates['hsd_rate'] = float(hsd_rate[0])
                    if power_rate and power_rate[0]:
                        procurement_rates['power_rate'] = float(power_rate[0])
                    
                    if procurement_rates:
                        summary['procurement_fuel_rates'] = procurement_rates
                        summary['procurement_fuel_rates']['source'] = 'procurement_data'
                except Exception as e:
                    logger.error(f"Error getting procurement fuel rates: {str(e)}")
            except Exception as e:
                logger.error(f"Error getting fuel rates: {str(e)}")
                
            # Add HPCL credit repayment calculation data
            try:
                # Get most recent total outstanding amount
                latest_outstanding = latest_entry.total_outstanding
                summary['hpcl_credit'] = {
                    'current_outstanding': float(latest_outstanding) if latest_outstanding else 0.0,
                    'as_of_date': latest_entry.date.isoformat()
                }
                
                # Calculate average daily collection over last 7 days for more accurate repayment estimate
                recent_collections = DailyConsolidation.query.order_by(DailyConsolidation.date.desc()).limit(7).all()
                if recent_collections:
                    # Calculate how many actual days we have data for (up to 7)
                    collection_days = len(recent_collections)
                    
                    # Calculate total collections across all days
                    total_collections = sum(
                        entry.cash_collections + entry.card_collections + entry.paytm_collections + entry.hp_transactions
                        for entry in recent_collections
                    )
                    
                    # Calculate average daily collection
                    avg_daily_collection = total_collections / collection_days if collection_days > 0 else 0
                    
                    # Also store the most recent collection for reference
                    latest_collection = (
                        recent_collections[0].cash_collections + 
                        recent_collections[0].card_collections + 
                        recent_collections[0].paytm_collections + 
                        recent_collections[0].hp_transactions
                    ) if recent_collections else 0
                    
                    # Store both values
                    summary['hpcl_credit']['daily_repayment_capacity'] = float(avg_daily_collection)
                    summary['hpcl_credit']['latest_collection'] = float(latest_collection)
                    summary['hpcl_credit']['avg_collection_days'] = collection_days
                    
                    # Calculate days based on average
                    summary['hpcl_credit']['repayment_days'] = int(float(latest_outstanding) / avg_daily_collection) if avg_daily_collection > 0 else 0
                    
            except Exception as e:
                logger.error(f"Error calculating HPCL credit repayment data: {str(e)}")
        
        # Calculate average daily consumption for last 7, 14, and 30 days
        consumption_stats = {}
        for days in [7, 14, 30]:
            if (newest_date - oldest_date).days >= days:
                cutoff_date = today - timedelta(days=days)
                recent_entries = DailyConsolidation.query.filter(DailyConsolidation.date >= cutoff_date).all()
                
                if recent_entries:
                    total_days = min(days, (newest_date - cutoff_date).days + 1)
                    
                    ms_quantity = sum(entry.ms_quantity for entry in recent_entries)
                    hsd_quantity = sum(entry.hsd_quantity for entry in recent_entries)
                    power_quantity = sum(entry.power_quantity for entry in recent_entries)
                    
                    consumption_stats[f'{days}_day_avg'] = {
                        'ms_daily_avg': ms_quantity / total_days if total_days > 0 else 0,
                        'hsd_daily_avg': hsd_quantity / total_days if total_days > 0 else 0,
                        'power_daily_avg': power_quantity / total_days if total_days > 0 else 0,
                        'actual_days_used': total_days,
                        'date_range': {
                            'from': cutoff_date.isoformat(),
                            'to': newest_date.isoformat()
                        }
                    }
        
        if consumption_stats:
            summary['consumption_statistics'] = consumption_stats
        
        # Get daily collection breakdown for analysis
        collection_data = db.session.query(
            DailyConsolidation.date,
            func.sum(DailyConsolidation.cash_collections).label('cash'),
            func.sum(DailyConsolidation.card_collections).label('card'),
            func.sum(DailyConsolidation.paytm_collections).label('paytm'),
            func.sum(DailyConsolidation.hp_transactions).label('hp')
        ).group_by(DailyConsolidation.date).order_by(DailyConsolidation.date.desc()).limit(30).all()
        
        if collection_data:
            summary['recent_collections'] = [
                {
                    'date': entry[0].isoformat(),
                    'cash': float(entry[1]) if entry[1] else 0,
                    'card': float(entry[2]) if entry[2] else 0,
                    'paytm': float(entry[3]) if entry[3] else 0,
                    'hp': float(entry[4]) if entry[4] else 0,
                    'total': float(sum(e for e in entry[1:5] if e))
                } for entry in collection_data
            ]
            
        # Get HPCL credit trend
        credit_trend = db.session.query(
            DailyConsolidation.date,
            DailyConsolidation.total_outstanding
        ).order_by(DailyConsolidation.date.desc()).limit(30).all()
        
        if credit_trend:
            summary['hpcl_credit_trend'] = [
                {
                    'date': entry[0].isoformat(),
                    'amount': float(entry[1]) if entry[1] else 0
                } for entry in credit_trend
            ]
            
        # Sales by day of week
        sales_by_day = db.session.query(
            func.strftime('%w', DailyConsolidation.date).label('day_of_week'),
            func.sum(DailyConsolidation.ms_amount + DailyConsolidation.hsd_amount + DailyConsolidation.power_amount).label('total_sales')
        ).group_by('day_of_week').all()
        
        if sales_by_day:
            day_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
            summary['sales_by_day_of_week'] = {
                day_names[int(day)]: float(amount) for day, amount in sales_by_day if amount
            }
            
        # Sales by shift
        sales_by_shift = db.session.query(
            DailyConsolidation.shift,
            func.sum(DailyConsolidation.ms_amount + DailyConsolidation.hsd_amount + DailyConsolidation.power_amount).label('total_sales'),
            func.count(DailyConsolidation.id).label('count')
        ).group_by(DailyConsolidation.shift).all()
        
        if sales_by_shift:
            summary['sales_by_shift'] = {
                shift: {
                    'total_sales': float(amount),
                    'count': count,
                    'average_per_shift': float(amount) / count if count else 0
                } for shift, amount, count in sales_by_shift
            }

    # Procurement entries summary
    procurement_entries_count = db.session.query(ProcurementData).count()
    if procurement_entries_count > 0:
        summary['procurement_data_summary'] = {
            'total_entries': procurement_entries_count,
            'date_range': {
                'oldest': db.session.query(func.min(ProcurementData.invoice_date)).scalar().isoformat(),
                'newest': db.session.query(func.max(ProcurementData.invoice_date)).scalar().isoformat()
            },
            'total_procurement_amount': db.session.query(func.sum(ProcurementData.total_amount)).scalar(),
            'total_ms_quantity_procured': db.session.query(func.sum(ProcurementData.quantity)).filter(ProcurementData.fuel_type == 'MS').scalar(),
            'total_hsd_quantity_procured': db.session.query(func.sum(ProcurementData.quantity)).filter(ProcurementData.fuel_type == 'HSD').scalar(),
            'total_power_quantity_procured': db.session.query(func.sum(ProcurementData.quantity)).filter(ProcurementData.fuel_type == 'POWER').scalar(),
        }
        
        # Recent procurement entries
        recent_procurements = ProcurementData.query.order_by(ProcurementData.invoice_date.desc()).limit(10).all()
        if recent_procurements:
            summary['recent_procurements'] = [
                {
                    'date': entry.invoice_date.isoformat(),
                    'fuel_type': entry.fuel_type,
                    'quantity': entry.quantity,
                    'rate': entry.rate,
                    'amount': entry.total_amount
                } for entry in recent_procurements
            ]
            
        # Latest fuel rates
        latest_rates = {}
        for fuel_type in ['MS', 'HSD', 'POWER']:
            latest = ProcurementData.query.filter_by(fuel_type=fuel_type).order_by(ProcurementData.invoice_date.desc()).first()
            if latest:
                latest_rates[fuel_type] = {
                    'rate': latest.rate,
                    'as_of_date': latest.invoice_date.isoformat()
                }
        
        if latest_rates:
            summary['latest_fuel_rates'] = latest_rates

    return summary

def get_daily_completion_status(target_date):
    """Check if daily entries are complete for a given date"""
    entries = DailyConsolidation.query.filter_by(date=target_date).all()
    
    morning_complete = any(entry.shift == 'Day' for entry in entries)
    night_complete = any(entry.shift == 'Night' for entry in entries)
    
    if morning_complete and night_complete:
        return 'COMPLETE'
    elif morning_complete or night_complete:
        return 'PARTIAL'
    else:
        return 'EMPTY'

# LangChain SQL Agent Functions
def create_sql_agent_instance():
    """Create and configure the LangChain SQL agent"""
    if not LANGCHAIN_AVAILABLE:
        raise ImportError("LangChain is not available. Please install required packages.")
    
    if not openai_api_key:
        raise ValueError("OpenAI API key is not configured")
    
    try:
        # Create the database connection for LangChain
        # Use the same database URI as the Flask app
        database_uri = app.config['SQLALCHEMY_DATABASE_URI']
        db_langchain = SQLDatabase.from_uri(database_uri)
        
        # Create ChatOpenAI instance
        llm = ChatOpenAI(
            model="gpt-4",
            temperature=0,
            openai_api_key=openai_api_key
        )
        
        # Create SQL toolkit
        toolkit = SQLDatabaseToolkit(db=db_langchain, llm=llm)
        
        # Create the SQL agent with safety configurations
        agent_executor = create_sql_agent(
            llm=llm,
            toolkit=toolkit,
            agent_type=AgentType.ZERO_SHOT_REACT_DESCRIPTION,
            verbose=True,
            handle_parsing_errors=True,
            max_iterations=3,
            early_stopping_method="generate"
        )
        
        return agent_executor, db_langchain
        
    except Exception as e:
        logger.error(f"Error creating SQL agent: {str(e)}")
        raise

def validate_sql_query(query):
    """Validate that SQL query is safe (read-only)"""
    # Convert to lowercase for checking
    query_lower = query.lower().strip()
    
    # Check for forbidden operations
    forbidden_keywords = [
        'insert', 'update', 'delete', 'drop', 'create', 'alter', 
        'truncate', 'replace', 'merge', 'call', 'exec', 'execute',
        'grant', 'revoke', 'commit', 'rollback', 'savepoint'
    ]
    
    for keyword in forbidden_keywords:
        if keyword in query_lower:
            return False, f"Forbidden operation '{keyword}' detected"
    
    # Must start with SELECT
    if not query_lower.startswith('select'):
        return False, "Only SELECT statements are allowed"
    
    return True, "Query is safe"

def get_database_schema_info():
    """Get schema information for the AI agent"""
    return """
    DATABASE SCHEMA INFORMATION:
    
    Table: daily_consolidation
    - Primary table for daily operations data
    - Columns: id, date, shift (Day/Night), manager, ms_rate, ms_quantity, ms_amount, 
              hsd_rate, hsd_quantity, hsd_amount, power_rate, power_quantity, power_amount,
              hsd1_tank, hsd2_tank, ms1_tank, ms2_tank, power1_tank, total_outstanding, hpcl_payment,
              cash_collections, card_collections, paytm_collections, hp_transactions, manager_notes, created_at
    
    Table: procurement_data
    - Fuel procurement/purchase records
    - Columns: id, invoice_number, invoice_date, fuel_type (MS/HSD/POWER), quantity, rate, 
              total_amount, vehicle_number, supplier, created_at
    
    Table: chat_history
    - AI chat conversation history
    - Columns: id, user_message, ai_response, timestamp
    
    BUSINESS CONTEXT:
    - This is a petrol station management system
    - Fuel types: MS (Motor Spirit/Petrol), HSD (High Speed Diesel), POWER (Power Petrol)
    - Shifts: Day (morning), Night (evening)
    - Collections: cash_collections, card_collections, paytm_collections, hp_transactions
    - Tank levels are measured in liters
    - HPCL is the fuel supplier, total_outstanding is credit owed to them
    - hpcl_payment tracks daily payments made to HPCL
    """

# API Routes

# Test route
@app.route('/api/test', methods=['GET'])
def test_api():
    return jsonify({'status': 'success', 'message': 'API is working'})

# Daily Consolidation Routes
@app.route('/api/daily-consolidation', methods=['POST'])
def create_daily_entry():
    try:
        data = request.json
        logger.debug(f"Received data: {data}")
        
        # Debug log specifically for HPCL value
        total_outstanding_value = data.get('total_outstanding', 0)
        logger.debug(f"HPCL Credit total_outstanding received: {total_outstanding_value}, type: {type(total_outstanding_value)}")
        
        # Convert date string to date object - this is the business day
        business_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        logger.debug(f"Business date: {business_date}")
        
        # Get current time for tracking when entry was made
        entry_timestamp = datetime.now()
        
        # Determine if this is a night shift entry
        shift = data['shift']
        
        # Add business day validation and logging
        current_business_day = get_current_business_day()
        logger.info(f"Creating entry for business date: {business_date}, current business day: {current_business_day}")
        
        if shift == 'Night':
            logger.info(f"Night shift entry for business day {business_date} - data covers {business_date} 8:30 PM to {business_date + timedelta(days=1)} 8:30 AM")
        else:
            logger.info(f"Day shift entry for business day {business_date} - data covers {business_date} 8:30 AM to {business_date} 8:30 PM")
        
        # Create entry with proper type conversion
        entry = DailyConsolidation(
            date=business_date,  # This is the business day date
            shift=shift,
            manager=data['manager'],
            ms_rate=float(data.get('ms_rate', 0)),
            ms_quantity=float(data.get('ms_quantity', 0)),
            ms_amount=float(data.get('ms_amount', 0)),
            hsd_rate=float(data.get('hsd_rate', 0)),
            hsd_quantity=float(data.get('hsd_quantity', 0)),
            hsd_amount=float(data.get('hsd_amount', 0)),
            power_rate=float(data.get('power_rate', 0)),
            power_quantity=float(data.get('power_quantity', 0)),
            power_amount=float(data.get('power_amount', 0)),
            hsd1_tank=float(data.get('hsd1_tank', 0)),
            hsd2_tank=float(data.get('hsd2_tank', 0)),
            ms1_tank=float(data.get('ms1_tank', 0)),
            ms2_tank=float(data.get('ms2_tank', 0)),
            power1_tank=float(data.get('power1_tank', 0)),
            total_outstanding=float(data.get('total_outstanding', 0)),
            hpcl_payment=float(data.get('hpcl_payment', 0)),
            cash_collections=float(data.get('cash_collections', 0)),
            card_collections=float(data.get('card_collections', 0)),
            paytm_collections=float(data.get('paytm_collections', 0)),
            hp_transactions=float(data.get('hp_transactions', 0)),
            manager_notes=data.get('manager_notes', '')
        )
        
        logger.debug("Created entry object")
        logger.debug(f"HPCL Credit total_outstanding in object: {entry.total_outstanding}")
        db.session.add(entry)
        db.session.commit()
        
        # Create or update tank reading for this business date
        try:
            # Check if a tank reading exists for this business date
            existing_reading = TankReading.query.filter_by(date=business_date).first()
            
            if existing_reading:
                # Update existing reading with latest values
                existing_reading.hsd1_tank = float(data.get('hsd1_tank', 0))
                existing_reading.hsd2_tank = float(data.get('hsd2_tank', 0))
                existing_reading.ms1_tank = float(data.get('ms1_tank', 0))
                existing_reading.ms2_tank = float(data.get('ms2_tank', 0))
                existing_reading.power1_tank = float(data.get('power1_tank', 0))
                # Update notes to reflect both shifts if needed
                if shift == 'Night':
                    existing_reading.notes = f"Updated from {shift} shift entry (Business Day: {business_date})"
                db.session.commit()
                logger.info(f"Updated tank reading for business date: {business_date}")
            else:
                # Create new tank reading for this business date
                reading = TankReading(
                    date=business_date,
                    hsd1_tank=float(data.get('hsd1_tank', 0)),
                    hsd2_tank=float(data.get('hsd2_tank', 0)),
                    ms1_tank=float(data.get('ms1_tank', 0)),
                    ms2_tank=float(data.get('ms2_tank', 0)),
                    power1_tank=float(data.get('power1_tank', 0)),
                    notes=f"Created from {shift} shift entry (Business Day: {business_date})"
                )
                db.session.add(reading)
                db.session.commit()
                logger.info(f"Created new tank reading for date: {business_date}")
        except Exception as e:
            logger.error(f"Error updating tank readings: {str(e)}")
            # Don't fail the whole transaction if tank reading update fails
        
        logger.info(f"Daily entry saved successfully - Date: {business_date}, Outstanding: {entry.total_outstanding}")
        
        return jsonify({'success': True, 'message': 'Daily entry created successfully'})
    except Exception as e:
        logger.error(f"Error creating daily entry: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/daily-consolidation', methods=['GET'])
def get_daily_entries():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = DailyConsolidation.query
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(DailyConsolidation.date >= start_date_obj)
        
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(DailyConsolidation.date <= end_date_obj)
        
        entries = query.order_by(DailyConsolidation.date.desc(), DailyConsolidation.created_at.desc()).all()
        
        logger.debug(f"Retrieved {len(entries)} daily entries")
        
        return jsonify({
            'success': True,
            'data': [entry.to_dict() for entry in entries]
        })
    except Exception as e:
        logger.error(f"Error retrieving daily entries: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/daily-consolidation/<int:entry_id>', methods=['PUT'])
def update_daily_entry(entry_id):
    try:
        entry = DailyConsolidation.query.get_or_404(entry_id)
        data = request.json
        
        # Handle date update if provided
        if 'date' in data and data['date']:
            try:
                # Parse the date
                if isinstance(data['date'], str):
                    # Handle different date formats
                    if 'T' in data['date']:
                        new_date = datetime.fromisoformat(data['date'].replace('Z', '+00:00')).date()
                    else:
                        new_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
                else:
                    new_date = data['date']
                
                entry.date = new_date
                logger.info(f"Updated entry {entry_id} date to: {new_date}")
            except (ValueError, TypeError) as e:
                logger.error(f"Error parsing date {data['date']}: {e}")
                return jsonify({'success': False, 'error': f'Invalid date format: {data["date"]}'}), 400
        
        # Update fields
        entry.shift = data.get('shift', entry.shift)
        entry.manager = data.get('manager', entry.manager)
        entry.ms_rate = float(data.get('ms_rate', entry.ms_rate))
        entry.ms_quantity = float(data.get('ms_quantity', entry.ms_quantity))
        entry.ms_amount = float(data.get('ms_amount', entry.ms_amount))
        entry.hsd_rate = float(data.get('hsd_rate', entry.hsd_rate))
        entry.hsd_quantity = float(data.get('hsd_quantity', entry.hsd_quantity))
        entry.hsd_amount = float(data.get('hsd_amount', entry.hsd_amount))
        entry.power_rate = float(data.get('power_rate', entry.power_rate))
        entry.power_quantity = float(data.get('power_quantity', entry.power_quantity))
        entry.power_amount = float(data.get('power_amount', entry.power_amount))
        entry.hsd1_tank = float(data.get('hsd1_tank', entry.hsd1_tank))
        entry.hsd2_tank = float(data.get('hsd2_tank', entry.hsd2_tank))
        entry.ms1_tank = float(data.get('ms1_tank', entry.ms1_tank))
        entry.ms2_tank = float(data.get('ms2_tank', entry.ms2_tank))
        entry.power1_tank = float(data.get('power1_tank', entry.power1_tank))
        entry.total_outstanding = float(data.get('total_outstanding', entry.total_outstanding))
        entry.hpcl_payment = float(data.get('hpcl_payment', entry.hpcl_payment))
        entry.cash_collections = float(data.get('cash_collections', entry.cash_collections))
        entry.card_collections = float(data.get('card_collections', entry.card_collections))
        entry.paytm_collections = float(data.get('paytm_collections', entry.paytm_collections))
        entry.hp_transactions = float(data.get('hp_transactions', entry.hp_transactions))
        entry.manager_notes = data.get('manager_notes', entry.manager_notes)
        
        db.session.commit()
        logger.info(f"Updated daily entry {entry_id}")
        
        # Update tank reading for this date
        try:
            # Check if a tank reading exists for this date
            existing_reading = TankReading.query.filter_by(date=entry.date).first()
            
            if existing_reading:
                # Update existing reading
                existing_reading.hsd1_tank = entry.hsd1_tank
                existing_reading.hsd2_tank = entry.hsd2_tank
                existing_reading.ms1_tank = entry.ms1_tank
                existing_reading.ms2_tank = entry.ms2_tank
                existing_reading.power1_tank = entry.power1_tank
                db.session.commit()
                logger.info(f"Updated tank reading for date: {entry.date}")
            else:
                # Create new tank reading
                reading = TankReading(
                    date=entry.date,
                    hsd1_tank=entry.hsd1_tank,
                    hsd2_tank=entry.hsd2_tank,
                    ms1_tank=entry.ms1_tank,
                    ms2_tank=entry.ms2_tank,
                    power1_tank=entry.power1_tank,
                    notes=f"Automatically created from updated daily entry - {entry.shift} shift"
                )
                db.session.add(reading)
                db.session.commit()
                logger.info(f"Created new tank reading for date: {entry.date}")
        except Exception as e:
            logger.error(f"Error updating tank readings: {str(e)}")
            # Don't fail the whole transaction if tank reading update fails
        
        return jsonify({'success': True, 'message': 'Daily entry updated successfully'})
    except Exception as e:
        logger.error(f"Error updating daily entry: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/daily-consolidation/<int:entry_id>', methods=['DELETE'])
def delete_daily_entry(entry_id):
    try:
        entry = DailyConsolidation.query.get_or_404(entry_id)
        db.session.delete(entry)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Daily entry deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# HPCL Ledger Route
@app.route('/api/hpcl-ledger', methods=['GET'])
def get_hpcl_ledger():
    try:
        # Log that the API was called
        logger.info("HPCL Ledger API called")
        
        # Get optional date range parameters
        days = request.args.get('days', default=30, type=int)
        logger.info(f"Requested days: {days}")
        
        # Calculate the date range
        today = date.today()
        start_date = today - timedelta(days=days)
        logger.info(f"Date range: {start_date} to {today}")
        
        # First check if the table exists
        try:
            # Try a simple query to check if the table exists
            db.session.query(DailyConsolidation).limit(1).all()
            logger.info("DailyConsolidation table exists")
        except Exception as e:
            logger.error(f"Table check error: {str(e)}")
            return jsonify({"success": False, "message": "Database table not found"}), 500
            
        # Check if hpcl_payment column exists in the database schema
        column_exists = False
        try:
            # Check if the column exists in the database using raw SQL
            with db.engine.connect() as conn:
                result = conn.execute(text("PRAGMA table_info(daily_consolidation)"))
                columns = [row[1] for row in result]
                column_exists = 'hpcl_payment' in columns
                logger.info(f"Column check via SQL: hpcl_payment {'exists' if column_exists else 'does not exist'}")
        except Exception as e:
            logger.warning(f"Error checking for column via SQL: {str(e)}")
            # Continue with the fallback check
        
        # Also check if the column exists in the SQLAlchemy model as a fallback
        if not column_exists:
            try:
                # This line will throw an exception if the column doesn't exist in the model
                getattr(DailyConsolidation, 'hpcl_payment')
                column_exists = True
                logger.info("hpcl_payment column exists in the database model")
            except AttributeError:
                logger.warning("hpcl_payment column does not exist in the database model")
        
        logger.info(f"Final column check result: hpcl_payment column {'exists' if column_exists else 'does not exist'}")
        
        # Query the database for daily entries within the date range
        try:
            query = db.session.query(
                DailyConsolidation.date,
                func.sum(DailyConsolidation.ms_amount + DailyConsolidation.hsd_amount + DailyConsolidation.power_amount).label('total_sales'),
                func.sum(DailyConsolidation.cash_collections + DailyConsolidation.card_collections + 
                         DailyConsolidation.paytm_collections + DailyConsolidation.hp_transactions).label('total_collections'),
                # Get the last total_outstanding value for each date
                func.max(DailyConsolidation.total_outstanding).label('total_outstanding')
            )
            
            # Conditionally add the hpcl_payment column to the query if it exists
            if column_exists:
                query = query.add_columns(func.sum(DailyConsolidation.hpcl_payment).label('hpcl_payment'))
            
            # Complete the query
            entries = query.filter(
                DailyConsolidation.date >= start_date
            ).group_by(
                DailyConsolidation.date
            ).order_by(
                DailyConsolidation.date.desc()
            ).all()
            
            logger.info(f"Query executed successfully. Found {len(entries)} entries.")
        except Exception as e:
            logger.error(f"Database query error: {str(e)}")
            return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500
        
        # Format the results
        ledger_data = []
        for entry in entries:
            entry_data = {
                'date': entry.date.isoformat(),
                'total_sales': float(entry.total_sales) if entry.total_sales else 0,
                'total_collections': float(entry.total_collections) if entry.total_collections else 0,
                'total_outstanding': float(entry.total_outstanding) if entry.total_outstanding else 0
            }
            
            # Add hpcl_payment if the column exists, otherwise default to 0
            if column_exists:
                entry_data['hpcl_payment'] = float(entry.hpcl_payment) if entry.hpcl_payment else 0
            else:
                entry_data['hpcl_payment'] = 0
            
            ledger_data.append(entry_data)
        
        return jsonify({
            'success': True,
            'ledger': ledger_data
        })
    except Exception as e:
        error_msg = f"Error retrieving HPCL ledger data: {str(e)}"
        logger.error(error_msg)
        print(error_msg)  # Print to console for immediate visibility
        
        # Return a more detailed error
        return jsonify({
            'success': False, 
            'error': str(e),
            'message': 'An error occurred while retrieving HPCL ledger data. The database might need to be updated with the new column.',
            'details': error_msg
        }), 400

# New HPCL Transaction Ledger Route
@app.route('/api/hpcl-transaction-ledger', methods=['GET'])
def get_hpcl_transaction_ledger():
    try:
        logger.info("HPCL Transaction Ledger API called - SHOWING PROCUREMENT, PAYMENTS, AND MANUAL BALANCES")
        
        # Get optional date range parameters
        days = request.args.get('days', default=30, type=int)
        logger.info(f"Requested days: {days}")
        
        # Calculate the date range
        today = date.today()
        start_date = today - timedelta(days=days)
        logger.info(f"Date range: {start_date} to {today}")
        
        # Collect all transactions
        transactions = []
        
        # 1. Get procurement transactions (these INCREASE credit/debt - Debit entries)
        try:
            procurement_entries = db.session.query(ProcurementData).filter(
                ProcurementData.invoice_date >= start_date
            ).order_by(ProcurementData.invoice_date.desc()).all()
            
            for entry in procurement_entries:
                # Get manual balance for this date (if any)
                manual_balance_entry = db.session.query(DailyConsolidation.total_outstanding).filter(
                    DailyConsolidation.date == entry.invoice_date,
                    DailyConsolidation.total_outstanding.isnot(None)
                ).first()
                
                manual_balance = float(manual_balance_entry.total_outstanding) if manual_balance_entry and manual_balance_entry.total_outstanding is not None else None
                
                # DEBUG: Check if this is affecting August 6th
                if entry.invoice_date.strftime('%Y-%m-%d') == '2025-08-06':
                    logger.error(f"AUGUST 6TH PROCUREMENT DEBUG - Invoice {entry.invoice_number}: Manual balance lookup = {manual_balance}")
                
                transactions.append({
                    'date': entry.invoice_date,
                    'description': f"Procurement - Invoice #{entry.invoice_number} ({entry.fuel_type} - {entry.quantity}L)",
                    'debit': float(entry.total_amount),  # Increases credit outstanding
                    'credit': 0,
                    'type': 'procurement',
                    'reference': entry.invoice_number,
                    'details': {
                        'fuel_type': entry.fuel_type,
                        'quantity': entry.quantity,
                        'rate': entry.rate,
                        'vehicle_number': entry.vehicle_number
                    },
                    'reported_balance': manual_balance,
                    'total_collections': 0
                })
            logger.info(f"Found {len(procurement_entries)} procurement transactions")
        except Exception as e:
            logger.error(f"Error fetching procurement data: {str(e)}")
        
        # 2. Get payment transactions (these DECREASE credit/debt - Credit entries)
        try:
            # Check if hpcl_payment column exists
            column_exists = False
            try:
                with db.engine.connect() as conn:
                    result = conn.execute(text("PRAGMA table_info(daily_consolidation)"))
                    columns = [row[1] for row in result]
                    column_exists = 'hpcl_payment' in columns
            except Exception as e:
                logger.warning(f"Error checking for hpcl_payment column: {str(e)}")
            
            if column_exists:
                payment_entries = db.session.query(DailyConsolidation).filter(
                    DailyConsolidation.date >= start_date,
                    DailyConsolidation.hpcl_payment > 0
                ).order_by(DailyConsolidation.date.desc()).all()
                
                for entry in payment_entries:
                    # Use the manual balance from this same entry
                    manual_balance = float(entry.total_outstanding) if entry.total_outstanding is not None else None
                    
                    # DEBUG: Check if this is affecting August 6th
                    if entry.date.strftime('%Y-%m-%d') == '2025-08-06':
                        logger.error(f"AUGUST 6TH PAYMENT DEBUG - Payment entry: Manual balance = {manual_balance}")
                    
                    transactions.append({
                        'date': entry.date,
                        'description': f"Payment to HPCL",
                        'debit': 0,
                        'credit': float(entry.hpcl_payment),  # Decreases credit outstanding
                        'type': 'payment',
                        'reference': f"Payment-{entry.date.strftime('%Y%m%d')}",
                        'details': {},
                        'reported_balance': manual_balance,
                        'total_collections': 0
                    })
                logger.info(f"Found {len(payment_entries)} payment transactions")
            else:
                logger.warning("hpcl_payment column not found - no payment transactions available")
        except Exception as e:
            logger.error(f"Error fetching payment data: {str(e)}")
        
        # 3. Get individual daily consolidation entries to show shift-level collections with EXACT manual balances
        try:
            # Get all daily consolidation entries in the date range
            daily_entries = db.session.query(DailyConsolidation).filter(
                DailyConsolidation.date >= start_date
            ).order_by(DailyConsolidation.date.desc(), DailyConsolidation.shift.desc()).all()
            
            logger.info(f"Found {len(daily_entries)} daily entries")
            
            # Add daily consolidation entries to show individual shift collections with EXACT manual balances
            for entry in daily_entries:
                # Calculate total collections for this specific shift
                cash_amount = float(entry.cash_collections or 0)
                card_amount = float(entry.card_collections or 0)
                paytm_amount = float(entry.paytm_collections or 0)
                hp_amount = float(entry.hp_transactions or 0)
                total_collections = cash_amount + card_amount + paytm_amount + hp_amount
                
                shift_display = "Day Shift (8:30 AM - 8:30 PM)" if entry.shift == "Day" else "Night Shift (8:30 PM - 8:30 AM)"
                
                # THE EXACT MANUAL BALANCE - NO CALCULATIONS!
                manual_balance = float(entry.total_outstanding) if entry.total_outstanding is not None else None
                
                logger.debug(f"Entry {entry.id}: Date={entry.date}, Shift={entry.shift}, Manual Balance={manual_balance}")
                
                # EXTRA DEBUG: Print the exact value for August 6th
                if entry.date.strftime('%Y-%m-%d') == '2025-08-06':
                    logger.error(f"AUGUST 6TH DEBUG - Entry ID {entry.id}: Manual outstanding = {entry.total_outstanding}, type = {type(entry.total_outstanding)}")
                    logger.error(f"AUGUST 6TH DEBUG - Calculated manual_balance = {manual_balance}")
                
                transactions.append({
                    'date': entry.date,
                    'description': f"Daily Collections - {shift_display} (Manager: {entry.manager or 'N/A'})",
                    'debit': 0,
                    'credit': 0,
                    'type': 'daily_collections',
                    'reference': f"Daily-{entry.date.strftime('%Y%m%d')}-{entry.shift}",
                    'details': {
                        'shift': entry.shift,
                        'manager': entry.manager,
                        'cash_collections': cash_amount,
                        'card_collections': card_amount,
                        'paytm_collections': paytm_amount,
                        'hp_transactions': hp_amount
                    },
                    'shift_collections': total_collections,
                    'shift_id': entry.id,
                    'reported_balance': manual_balance,  # EXACT manual value from data entry
                    'total_collections': total_collections
                })
            
            logger.info(f"Added {len(daily_entries)} individual shift collection entries with manual balances")
        except Exception as e:
            logger.error(f"Error fetching daily consolidation entries: {str(e)}")
        
        # Sort all transactions by date (newest first)
        transactions.sort(key=lambda x: x['date'], reverse=True)
        
        # Calculate key metrics for the period
        total_procured = sum(t['debit'] for t in transactions if t['type'] == 'procurement')
        total_payments = sum(t['credit'] for t in transactions if t['type'] == 'payment')
        net_change = total_procured - total_payments
        avg_daily_procurement = total_procured / days if days > 0 else 0
        
        # Format transactions for response - NO CALCULATIONS, JUST RAW MANUAL VALUES
        formatted_transactions = []
        for transaction in transactions:
            # Get the EXACT manual balance - no modifications
            reported_balance = transaction.get('reported_balance', None)
            
            formatted_transactions.append({
                'date': transaction['date'].isoformat(),
                'description': transaction['description'],
                'debit': float(transaction.get('debit', 0) or 0),
                'credit': float(transaction.get('credit', 0) or 0),
                'total_collections': float(transaction.get('total_collections', 0) or 0),
                'reported_balance': float(reported_balance) if reported_balance is not None else None,
                'type': transaction['type'],
                'reference': transaction['reference'],
                'details': transaction['details']
            })
        
        # Get latest reported balance for summary
        latest_reported = transactions[0]['reported_balance'] if transactions else None
        
        response_data = {
            'success': True,
            'transactions': formatted_transactions,
            'key_metrics': {
                'period_days': days,
                'total_procured': total_procured,
                'total_payments': total_payments,
                'net_change': net_change,
                'avg_daily_procurement': avg_daily_procurement,
                'latest_reported_balance': latest_reported
            },
            'summary': {
                'total_transactions': len(transactions),
                'procurement_transactions': sum(1 for t in transactions if t['type'] == 'procurement'),
                'payment_transactions': sum(1 for t in transactions if t['type'] == 'payment'),
                'daily_collection_entries': sum(1 for t in transactions if t['type'] == 'daily_collections')
            }
        }
        
        logger.info(f"Successfully generated transaction ledger with {len(transactions)} transactions showing procurement, payments, and manual balances")
        return jsonify(response_data)
        
    except Exception as e:
        error_msg = f"Error retrieving HPCL transaction ledger: {str(e)}"
        logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while retrieving HPCL transaction ledger data.'
        }), 400

# Customer Credit Routes
@app.route('/api/customer-credit', methods=['POST'])
def create_customer_credit():
    try:
        data = request.json
        
        # Convert date string to date object
        credit_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        entry = CustomerCredit(
            customer_name=data['customer_name'],
            date=credit_date,
            fuel_type=data['fuel_type'],
            quantity=float(data['quantity']),
            rate=float(data['rate']),
            total_amount=float(data['total_amount']),
            transaction_type=data['transaction_type'],  # 'sale' or 'payment'
            notes=data.get('notes', '')
        )
        
        db.session.add(entry)
        db.session.commit()
        
        logger.info(f"Customer credit {data['transaction_type']} created for {data['customer_name']}: {data['total_amount']}")
        
        return jsonify({
            'success': True,
            'message': f'Customer credit {data["transaction_type"]} recorded successfully.',
            'data': entry.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        error_msg = f"Error creating customer credit entry: {str(e)}"
        logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while creating the customer credit entry.'
        }), 400

@app.route('/api/customer-credit/overview', methods=['GET'])
def get_customer_credit_overview():
    try:
        # Get all customers and their outstanding balances
        customers = db.session.query(CustomerCredit.customer_name).distinct().all()
        customer_overviews = []
        
        for customer in customers:
            customer_name = customer[0]
            
            # Calculate outstanding balance for this customer
            sales = db.session.query(db.func.sum(CustomerCredit.total_amount)).filter(
                CustomerCredit.customer_name == customer_name,
                CustomerCredit.transaction_type == 'sale'
            ).scalar() or 0
            
            payments = db.session.query(db.func.sum(CustomerCredit.total_amount)).filter(
                CustomerCredit.customer_name == customer_name,
                CustomerCredit.transaction_type == 'payment'
            ).scalar() or 0
            
            outstanding_balance = float(sales) - float(payments)
            
            # Get last transaction date
            last_transaction = db.session.query(CustomerCredit.date).filter(
                CustomerCredit.customer_name == customer_name
            ).order_by(CustomerCredit.date.desc()).first()
            
            # Get transaction count
            transaction_count = db.session.query(CustomerCredit).filter(
                CustomerCredit.customer_name == customer_name
            ).count()
            
            customer_overviews.append({
                'customer_name': customer_name,
                'outstanding_balance': outstanding_balance,
                'total_sales': float(sales),
                'total_payments': float(payments),
                'last_transaction_date': last_transaction[0].isoformat() if last_transaction else None,
                'transaction_count': transaction_count
            })
        
        # Sort by outstanding balance (highest first)
        customer_overviews.sort(key=lambda x: x['outstanding_balance'], reverse=True)
        
        # Calculate summary metrics
        total_outstanding = sum(c['outstanding_balance'] for c in customer_overviews)
        total_customers = len(customer_overviews)
        customers_with_debt = sum(1 for c in customer_overviews if c['outstanding_balance'] > 0)
        
        return jsonify({
            'success': True,
            'customers': customer_overviews,
            'summary': {
                'total_outstanding': total_outstanding,
                'total_customers': total_customers,
                'customers_with_debt': customers_with_debt,
                'avg_outstanding_per_customer': total_outstanding / total_customers if total_customers > 0 else 0
            }
        })
        
    except Exception as e:
        error_msg = f"Error retrieving customer credit overview: {str(e)}"
        logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while retrieving customer credit overview.'
        }), 400

@app.route('/api/customer-credit/history/<customer_name>', methods=['GET'])
def get_customer_credit_history(customer_name):
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # Get all transactions for this customer
        transactions = db.session.query(CustomerCredit).filter(
            CustomerCredit.customer_name == customer_name
        ).order_by(CustomerCredit.date.desc(), CustomerCredit.id.desc()).all()
        
        # Calculate running balance for each transaction (working backwards)
        running_balance = 0
        for transaction in reversed(transactions):
            if transaction.transaction_type == 'sale':
                running_balance += transaction.total_amount
            else:  # payment
                running_balance -= transaction.total_amount
            transaction.running_balance = running_balance
        
        # Prepare transactions with running balance
        formatted_transactions = []
        for transaction in transactions:
            transaction_dict = transaction.to_dict()
            transaction_dict['running_balance'] = transaction.running_balance
            formatted_transactions.append(transaction_dict)
        
        # Calculate summary for this customer
        total_sales = sum(t.total_amount for t in transactions if t.transaction_type == 'sale')
        total_payments = sum(t.total_amount for t in transactions if t.transaction_type == 'payment')
        current_balance = total_sales - total_payments
        
        return jsonify({
            'success': True,
            'customer_name': customer_name,
            'transactions': formatted_transactions,
            'summary': {
                'total_sales': float(total_sales),
                'total_payments': float(total_payments),
                'current_balance': float(current_balance),
                'transaction_count': len(transactions),
                'first_transaction_date': transactions[-1].date.isoformat() if transactions else None,
                'last_transaction_date': transactions[0].date.isoformat() if transactions else None
            }
        })
        
    except Exception as e:
        error_msg = f"Error retrieving customer credit history for {customer_name}: {str(e)}"
        logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'An error occurred while retrieving credit history for {customer_name}.'
        }), 400

@app.route('/api/customer-credit/recent', methods=['GET'])
def get_recent_customer_credit():
    try:
        # Get recent transactions (last 30 days or last 100 transactions, whichever is smaller)
        days_back = request.args.get('days', 30, type=int)
        limit = request.args.get('limit', 100, type=int)
        
        start_date = datetime.now().date() - timedelta(days=days_back)
        
        transactions = db.session.query(CustomerCredit).filter(
            CustomerCredit.date >= start_date
        ).order_by(CustomerCredit.date.desc(), CustomerCredit.id.desc()).limit(limit).all()
        
        formatted_transactions = [transaction.to_dict() for transaction in transactions]
        
        # Calculate period summary
        period_sales = sum(t.total_amount for t in transactions if t.transaction_type == 'sale')
        period_payments = sum(t.total_amount for t in transactions if t.transaction_type == 'payment')
        
        return jsonify({
            'success': True,
            'transactions': formatted_transactions,
            'period_summary': {
                'days_covered': days_back,
                'transaction_count': len(transactions),
                'total_sales': float(period_sales),
                'total_payments': float(period_payments),
                'net_credit_change': float(period_sales - period_payments)
            }
        })
        
    except Exception as e:
        error_msg = f"Error retrieving recent customer credit transactions: {str(e)}"
        logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while retrieving recent customer credit transactions.'
        }), 400

# HPCL Payments Routes
@app.route('/api/hpcl-payments', methods=['GET'])
def get_hpcl_payments():
    try:
        # Get optional date range parameters
        days = request.args.get('days', default=30, type=int)
        
        # Calculate the date range
        today = date.today()
        start_date = today - timedelta(days=days)
        
        payments = HPCLPayments.query.filter(
            HPCLPayments.date >= start_date
        ).order_by(HPCLPayments.date.desc()).all()
        
        return jsonify({
            'success': True,
            'payments': [payment.to_dict() for payment in payments]
        })
        
    except Exception as e:
        logger.error(f"Error retrieving HPCL payments: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while retrieving HPCL payments.'
        }), 400

@app.route('/api/hpcl-payments', methods=['POST'])
def create_hpcl_payment():
    try:
        data = request.json
        
        # Convert date string to date object
        payment_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        payment = HPCLPayments(
            date=payment_date,
            amount=float(data['amount']),
            payment_method=data.get('payment_method', 'Bank Transfer'),
            reference_number=data.get('reference_number'),
            notes=data.get('notes')
        )
        
        db.session.add(payment)
        db.session.commit()
        
        logger.info(f"Created HPCL payment: ₹{payment.amount} on {payment.date}")
        
        return jsonify({
            'success': True, 
            'message': 'HPCL payment recorded successfully',
            'payment_id': payment.id
        })
        
    except Exception as e:
        logger.error(f"Error creating HPCL payment: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while recording the HPCL payment.'
        }), 400

@app.route('/api/hpcl-payments/<int:payment_id>', methods=['PUT'])
def update_hpcl_payment(payment_id):
    try:
        payment = HPCLPayments.query.get_or_404(payment_id)
        data = request.json
        
        # Update fields
        if 'date' in data:
            payment.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if 'amount' in data:
            payment.amount = float(data['amount'])
        if 'payment_method' in data:
            payment.payment_method = data['payment_method']
        if 'reference_number' in data:
            payment.reference_number = data['reference_number']
        if 'notes' in data:
            payment.notes = data['notes']
        
        db.session.commit()
        
        logger.info(f"Updated HPCL payment {payment_id}")
        
        return jsonify({
            'success': True,
            'message': 'HPCL payment updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Error updating HPCL payment {payment_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while updating the HPCL payment.'
        }), 400

@app.route('/api/hpcl-payments/<int:payment_id>', methods=['DELETE'])
def delete_hpcl_payment(payment_id):
    try:
        payment = HPCLPayments.query.get_or_404(payment_id)
        
        db.session.delete(payment)
        db.session.commit()
        
        logger.info(f"Deleted HPCL payment {payment_id}")
        
        return jsonify({
            'success': True,
            'message': 'HPCL payment deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting HPCL payment {payment_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'An error occurred while deleting the HPCL payment.'
        }), 400

# Procurement Routes
@app.route('/api/procurement', methods=['POST'])
def create_procurement_entry():
    try:
        data = request.json
        
        # Convert date string to date object
        invoice_date = datetime.strptime(data['invoice_date'], '%Y-%m-%d').date()
        
        entry = ProcurementData(
            invoice_number=data['invoice_number'],
            invoice_date=invoice_date,
            fuel_type=data['fuel_type'],
            quantity=float(data['quantity']),
            rate=float(data['rate']),
            total_amount=float(data['total_amount']),
            vehicle_number=data['vehicle_number'],
            supplier=data.get('supplier', 'HPCL')
        )
        
        db.session.add(entry)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Procurement entry created successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/procurement', methods=['GET'])
def get_procurement_entries():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = ProcurementData.query
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(ProcurementData.invoice_date >= start_date_obj)
        
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(ProcurementData.invoice_date <= end_date_obj)
        
        entries = query.order_by(ProcurementData.invoice_date.desc()).all()
        
        return jsonify({
            'success': True,
            'data': [entry.to_dict() for entry in entries]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/procurement/<int:entry_id>', methods=['PUT'])
def update_procurement_entry(entry_id):
    try:
        entry = ProcurementData.query.get_or_404(entry_id)
        data = request.json
        
        entry.invoice_number = data.get('invoice_number', entry.invoice_number)
        entry.fuel_type = data.get('fuel_type', entry.fuel_type)
        entry.quantity = float(data.get('quantity', entry.quantity))
        entry.rate = float(data.get('rate', entry.rate))
        entry.total_amount = float(data.get('total_amount', entry.total_amount))
        entry.vehicle_number = data.get('vehicle_number', entry.vehicle_number)
        entry.supplier = data.get('supplier', entry.supplier)
        
        if 'invoice_date' in data:
            entry.invoice_date = datetime.strptime(data['invoice_date'], '%Y-%m-%d').date()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Procurement entry updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/procurement/<int:entry_id>', methods=['DELETE'])
def delete_procurement_entry(entry_id):
    try:
        entry = ProcurementData.query.get_or_404(entry_id)
        db.session.delete(entry)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Procurement entry deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# Procurement Bulk Upload Routes
@app.route('/api/procurement/template', methods=['GET'])
def download_procurement_template():
    """Generate and download Excel template for procurement bulk upload"""
    try:
        # Create a sample data structure
        template_data = {
            'invoice_number': ['SAMPLE001', 'SAMPLE002'],
            'invoice_date': ['2024-01-01', '2024-01-01'],
            'fuel_type': ['HSD', 'MS'],
            'quantity': [1000.00, 500.00],
            'rate': [85.50, 92.30],
            'total_amount': [85500.00, 46150.00],
            'vehicle_number': ['MH12AB1234', 'MH12CD5678'],
            'supplier': ['HPCL', 'HPCL']
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the template with sample data
            df.to_excel(writer, sheet_name='Procurement_Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Procurement_Template']
            
            # Add instructions and formatting
            instructions = [
                "PROCUREMENT BULK UPLOAD TEMPLATE",
                "",
                "Instructions:",
                "1. Fill in all the columns with your procurement data",
                "2. invoice_date format: YYYY-MM-DD (e.g., 2024-01-01)",
                "3. fuel_type options: HSD, MS, POWER",
                "4. quantity should be in liters (decimal values allowed)",
                "5. rate should be in rupees per liter (decimal values allowed)", 
                "6. total_amount will be auto-calculated but you can override",
                "7. Delete the sample rows below before uploading",
                "",
                "Sample data (delete these rows):"
            ]
            
            # Insert instructions at the top
            worksheet.insert_rows(1, len(instructions))
            for i, instruction in enumerate(instructions, 1):
                worksheet[f'A{i}'] = instruction
            
            # Apply formatting
            from openpyxl.styles import Font, PatternFill
            
            # Header formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            # Format the actual header row
            header_row = len(instructions) + 1
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'procurement_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error generating procurement template: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/procurement/bulk-upload', methods=['POST'])
def bulk_upload_procurement():
    """Handle bulk procurement upload from Excel file"""
    try:
        logger.info("Bulk upload request received")
        
        if 'file' not in request.files:
            logger.error("No file in request")
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file
        try:
            excel_file = pd.ExcelFile(file)
            sheet_name = excel_file.sheet_names[0]
            
            # First, read the entire file to find headers
            temp_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            
            # Find the row containing column headers
            header_row = None
            for i in range(min(25, len(temp_df))):
                # Check if this row contains our expected column names
                row_values = [str(val).lower().strip() for val in temp_df.iloc[i].values if pd.notna(val)]
                
                # Look for key column names
                if any('invoice_number' in val for val in row_values) and any('fuel_type' in val for val in row_values):
                    header_row = i
                    break
                    
                # Alternative check - look for blue header row (often the data starts after instructions)
                if (len(row_values) >= 6 and 
                    'invoice' in ' '.join(row_values) and 
                    'date' in ' '.join(row_values) and
                    'fuel' in ' '.join(row_values)):
                    header_row = i
                    break
            
            if header_row is None:
                # Last resort - assume headers are in a reasonable position
                for i in range(5, min(20, len(temp_df))):
                    row_vals = temp_df.iloc[i].values
                    if sum(pd.notna(row_vals)) >= 6:  # At least 6 non-null values
                        header_row = i
                        break
            
            if header_row is None:
                raise ValueError("Could not find header row with column names. Please ensure your Excel file has the proper column headers.")
            
            logger.info(f"Found header row at index: {header_row}")
            
            # Now read with the correct header row
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
            
            # Remove any completely empty rows
            df = df.dropna(how='all')
            
            # Log the columns found
            logger.info(f"Columns after reading with header row {header_row}: {list(df.columns)}")
            
            if df is None or df.empty:
                raise ValueError("Could not find valid data in Excel file")
                
        except Exception as e:
            return jsonify({'success': False, 'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Log the actual columns found for debugging
        logger.info(f"Excel columns found: {list(df.columns)}")
        
        # Clean column names (remove extra spaces, convert to lowercase for comparison)
        df.columns = df.columns.str.strip().str.lower()
        
        # Validate required columns
        required_columns = ['invoice_number', 'invoice_date', 'fuel_type', 'quantity', 'rate', 'vehicle_number']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'error': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join(df.columns)}'
            }), 400
        
        # Process and validate data
        successful_entries = []
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row['invoice_number']) or str(row['invoice_number']).strip() == '':
                    continue
                
                # Skip sample data rows or instruction rows
                if str(row['invoice_number']).upper().startswith(('SAMPLE', 'INSTRUCTIONS', 'FILL')):
                    continue
                
                # Skip if it's a header row that got mixed in
                if str(row['invoice_number']).lower() == 'invoice_number':
                    continue
                
                # Validate and convert data
                invoice_date_str = str(row['invoice_date']).split()[0]  # Remove time part if exists
                try:
                    invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date()
                except:
                    try:
                        # Try alternative date format
                        invoice_date = pd.to_datetime(row['invoice_date']).date()
                    except:
                        raise ValueError(f"Invalid date format: {invoice_date_str}")
                
                fuel_type = str(row['fuel_type']).upper().strip()
                if fuel_type not in ['HSD', 'MS', 'POWER']:
                    raise ValueError(f"Invalid fuel type: {fuel_type}. Must be HSD, MS, or POWER")
                
                quantity = float(row['quantity'])
                if quantity <= 0:
                    raise ValueError("Quantity must be greater than 0")
                
                rate = float(row['rate'])
                if rate <= 0:
                    raise ValueError("Rate must be greater than 0")
                
                # Calculate total amount (use provided value if exists and valid, otherwise calculate)
                if 'total_amount' in df.columns and not pd.isna(row['total_amount']) and float(row['total_amount']) > 0:
                    total_amount = float(row['total_amount'])
                else:
                    total_amount = quantity * rate
                
                supplier = str(row.get('supplier', 'HPCL')).strip()
                if not supplier or supplier.upper() == 'NAN':
                    supplier = 'HPCL'
                
                # Create procurement entry
                entry = ProcurementData(
                    invoice_number=str(row['invoice_number']).strip(),
                    invoice_date=invoice_date,
                    fuel_type=fuel_type,
                    quantity=quantity,
                    rate=rate,
                    total_amount=total_amount,
                    vehicle_number=str(row['vehicle_number']).strip(),
                    supplier=supplier
                )
                
                db.session.add(entry)
                successful_entries.append({
                    'row': index + 1,
                    'invoice_number': entry.invoice_number,
                    'total_amount': entry.total_amount
                })
                
            except Exception as e:
                error_msg = f"Row {index + 1} (Invoice: {str(row.get('invoice_number', 'N/A'))}): {str(e)}"
                logger.error(error_msg)
                errors.append({
                    'row': index + 1,
                    'invoice_number': str(row.get('invoice_number', 'N/A')),
                    'error': str(e)
                })
        
        # Commit successful entries
        if successful_entries:
            db.session.commit()
            logger.info(f"Successfully uploaded {len(successful_entries)} procurement entries")
        else:
            db.session.rollback()
        
        return jsonify({
            'success': True,
            'message': f'Bulk upload completed. {len(successful_entries)} entries added successfully.',
            'successful_entries': len(successful_entries),
            'errors': len(errors),
            'error_details': errors if errors else None
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in bulk procurement upload: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Tank Readings Routes
@app.route('/api/tank-readings', methods=['GET', 'POST'])
def tank_readings():
    if request.method == 'GET':
        try:
            # Get date range filters if provided
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            
            query = TankReading.query
            
            if start_date and end_date:
                query = query.filter(
                    TankReading.date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
                    TankReading.date <= datetime.strptime(end_date, '%Y-%m-%d').date()
                )
                
            entries = query.order_by(TankReading.date.desc()).all()
            return jsonify({'success': True, 'data': [entry.to_dict() for entry in entries]})
        except Exception as e:
            logger.error(f"Error retrieving tank readings: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 400
            
    elif request.method == 'POST':
        try:
            data = request.json
            
            # Convert date string to date object
            reading_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
            
            # Check if reading already exists for this date
            existing_reading = TankReading.query.filter_by(date=reading_date).first()
            
            if existing_reading:
                # Update existing reading
                existing_reading.hsd1_tank = float(data.get('hsd1_tank', 0))
                existing_reading.hsd2_tank = float(data.get('hsd2_tank', 0))
                existing_reading.ms1_tank = float(data.get('ms1_tank', 0))
                existing_reading.ms2_tank = float(data.get('ms2_tank', 0))
                existing_reading.power1_tank = float(data.get('power1_tank', 0))
                existing_reading.notes = data.get('notes', '')
                
                db.session.commit()
                logger.info(f"Updated tank reading for date: {reading_date}")
                
                return jsonify({'success': True, 'message': 'Tank reading updated successfully', 'id': existing_reading.id})
            else:
                # Create new tank reading
                reading = TankReading(
                    date=reading_date,
                    hsd1_tank=float(data.get('hsd1_tank', 0)),
                    hsd2_tank=float(data.get('hsd2_tank', 0)),
                    ms1_tank=float(data.get('ms1_tank', 0)),
                    ms2_tank=float(data.get('ms2_tank', 0)),
                    power1_tank=float(data.get('power1_tank', 0)),
                    notes=data.get('notes', '')
                )
                
                db.session.add(reading)
                db.session.commit()
                logger.info(f"Tank reading saved for date: {reading_date}")
                
                return jsonify({'success': True, 'message': 'Tank reading saved successfully', 'id': reading.id})
        except Exception as e:
            logger.error(f"Error saving tank reading: {str(e)}")
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/tank-readings/<int:reading_id>', methods=['PUT'])
def update_tank_reading(reading_id):
    try:
        data = request.json
        reading = TankReading.query.get(reading_id)
        
        if not reading:
            return jsonify({'success': False, 'error': 'Tank reading not found'}), 404
            
        # Update fields
        if 'hsd1_tank' in data: reading.hsd1_tank = float(data['hsd1_tank'])
        if 'hsd2_tank' in data: reading.hsd2_tank = float(data['hsd2_tank'])
        if 'ms1_tank' in data: reading.ms1_tank = float(data['ms1_tank'])
        if 'ms2_tank' in data: reading.ms2_tank = float(data['ms2_tank'])
        if 'power1_tank' in data: reading.power1_tank = float(data['power1_tank'])
        if 'notes' in data: reading.notes = data['notes']
        
        db.session.commit()
        logger.info(f"Updated tank reading {reading_id}")
        
        return jsonify({'success': True, 'message': 'Tank reading updated successfully'})
    except Exception as e:
        logger.error(f"Error updating tank reading: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/tank-readings/<int:reading_id>', methods=['DELETE'])
def delete_tank_reading(reading_id):
    try:
        reading = TankReading.query.get(reading_id)
        
        if not reading:
            return jsonify({'success': False, 'error': 'Tank reading not found'}), 404
            
        db.session.delete(reading)
        db.session.commit()
        logger.info(f"Deleted tank reading {reading_id}")
        
        return jsonify({'success': True, 'message': 'Tank reading deleted successfully'})
    except Exception as e:
        logger.error(f"Error deleting tank reading: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

# AI Chat Routes
@app.route('/api/ai/chat', methods=['POST'])
def ai_chat():
    try:
        # Check if OpenAI API key is valid
        if not openai_api_key:
            logger.error("Invalid OpenAI API key. Please set a valid API key.")
            return jsonify({'success': False, 'error': "OpenAI API key is not configured correctly. Please set a valid API key."}), 400
            
        # Check if LangChain is available
        if not LANGCHAIN_AVAILABLE:
            logger.warning("LangChain is not available. Using fallback AI chat system.")
            # Fall back to basic AI chat with business data summary
            try:
                data = request.json
                if not data or 'message' not in data:
                    return jsonify({'success': False, 'error': "Message field is required"}), 400
                    
                user_message = data['message']
                logger.info(f"Processing fallback AI chat request: {user_message[:50]}...")
                
                # Get business data summary for context
                business_data = get_business_data_summary()
                
                # Create system prompt with business context
                system_prompt = f"""
                You are a helpful business analyst for a petrol station management system. 
                You have access to business data and can provide insights about fuel sales, inventory, collections, and operations.
                
                Business Data Summary:
                {json.dumps(business_data, indent=2)}
                
                Please provide helpful, data-driven responses based on the available information.
                Always format currency amounts in Indian Rupees (Rs.).
                """
                
                # Use OpenAI API directly
                if hasattr(openai, 'OpenAI'):
                    client = openai.OpenAI(api_key=openai_api_key)
                    response = client.chat.completions.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_message}
                        ]
                    )
                    ai_response = response.choices[0].message.content
                else:
                    # Legacy OpenAI client
                    response = openai.ChatCompletion.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_message}
                        ]
                    )
                    ai_response = response.choices[0].message.content
                
                # Save chat history
                chat_entry = ChatHistory(
                    user_message=user_message,
                    ai_response=f"[Fallback Mode] {ai_response}"
                )
                db.session.add(chat_entry)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'response': ai_response,
                    'mode': 'fallback'
                })
                
            except Exception as fallback_error:
                logger.error(f"Fallback AI chat failed: {str(fallback_error)}")
                return jsonify({'success': False, 'error': f"AI service temporarily unavailable: {str(fallback_error)}"}), 500
            
        data = request.json
        if not data:
            logger.error("No JSON data received in request")
            return jsonify({'success': False, 'error': "No data received in request"}), 400
            
        if 'message' not in data:
            logger.error("No 'message' field found in request data")
            return jsonify({'success': False, 'error': "Message field is required"}), 400
            
        user_message = data['message']
        logger.info(f"Processing LangChain SQL Agent request: {user_message[:50]}...")
        
        try:
            # Create database connection for LangChain
            database_uri = app.config['SQLALCHEMY_DATABASE_URI']
            db_langchain = SQLDatabase.from_uri(database_uri)
            
            # Create ChatOpenAI instance
            llm = ChatOpenAI(
                model="gpt-4",
                temperature=0,
                openai_api_key=openai_api_key
            )
            
            # Create SQL Database Toolkit
            toolkit = SQLDatabaseToolkit(db=db_langchain, llm=llm)
            
            # Create the SQL Agent with custom instructions
            agent_executor = create_sql_agent(
                llm=llm,
                toolkit=toolkit,
                verbose=True,
                agent_type=AgentType.ZERO_SHOT_REACT_DESCRIPTION,
                handle_parsing_errors=True,
                agent_executor_kwargs={
                    "return_intermediate_steps": True
                }
            )
            
            # Create enhanced system prompt for petrol station context
            enhanced_prompt = f"""
You are an expert business intelligence analyst for a petrol station management system. You have direct access to a SQLite database with the following key tables:

**Tables Available:**
1. **daily_consolidation** - Daily fuel sales, collections, tank levels, and manager information
2. **procurement_data** - Fuel procurement records with invoices, quantities, and rates
3. **chat_history** - Previous AI conversations

**Important Guidelines:**
1. **ONLY use SELECT statements** - Never INSERT, UPDATE, DELETE, or ALTER
2. **Always specify date ranges** when analyzing trends
3. **Calculate averages properly** - Use actual data points, not assumptions
4. **Currency format** - Always display amounts in Indian Rupees (Rs.)
5. **Be specific about timeframes** - State exact date ranges used in analysis
6. **Verify data quality** - Mention if data seems incomplete or inconsistent

**Key Business Metrics to Focus On:**
- Daily fuel sales (MS, HSD, POWER)
- Collection efficiency (cash, card, digital payments)
- Tank level monitoring and stock management
- HPCL credit outstanding and payment patterns
- Procurement vs. sales rate analysis

**When analyzing data:**
- Look for trends and anomalies
- Provide actionable insights
- Calculate realistic forecasts
- Compare different time periods
- Highlight any data quality issues

User Question: {user_message}

Please analyze the database to answer this question with specific data, calculations, and insights.
"""
            
            # Execute the query through LangChain SQL Agent
            logger.info("Executing LangChain SQL Agent...")
            result = agent_executor.invoke({"input": enhanced_prompt})
            
            # Extract the AI response
            ai_response = result.get("output", "I apologize, but I couldn't process your request properly.")
            
            # Log the intermediate steps for debugging
            if "intermediate_steps" in result:
                logger.info(f"SQL Agent executed {len(result['intermediate_steps'])} steps")
                for i, step in enumerate(result['intermediate_steps']):
                    logger.info(f"Step {i+1}: {step}")
            
            # Save chat history
            chat_entry = ChatHistory(
                user_message=user_message,
                ai_response=ai_response
            )
            db.session.add(chat_entry)
            db.session.commit()
            
            logger.info("LangChain SQL Agent request completed successfully")
            return jsonify({
                'success': True,
                'response': ai_response
            })
            
        except Exception as langchain_error:
            logger.error(f"LangChain SQL Agent error: {str(langchain_error)}")
            
            # Fallback to basic response if LangChain fails
            logger.info("Falling back to basic AI response...")
            
            try:
                # Basic fallback system message
                fallback_prompt = f"""
You are a helpful assistant for a petrol station management system. 
The user asked: {user_message}

Please provide a helpful response. Note that detailed data analysis is temporarily unavailable.
You can provide general guidance about petrol station operations, fuel management, and business best practices.
"""
                
                # Use basic OpenAI call as fallback
                if hasattr(openai, 'OpenAI'):
                    client = openai.OpenAI(api_key=openai_api_key)
                    response = client.chat.completions.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": fallback_prompt},
                            {"role": "user", "content": user_message}
                        ]
                    )
                    ai_response = response.choices[0].message.content
                else:
                    # Legacy OpenAI client
                    response = openai.ChatCompletion.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": fallback_prompt},
                            {"role": "user", "content": user_message}
                        ]
                    )
                    ai_response = response.choices[0].message.content
                
                # Save chat history
                chat_entry = ChatHistory(
                    user_message=user_message,
                    ai_response=f"[Fallback Mode] {ai_response}"
                )
                db.session.add(chat_entry)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'response': ai_response,
                    'fallback': True
                })
                
            except Exception as fallback_error:
                logger.error(f"Fallback AI chat also failed: {str(fallback_error)}")
                return jsonify({
                    'success': False, 
                    'error': f"Both LangChain and fallback AI systems failed: {str(langchain_error)}"
                }), 500
        
    except Exception as e:
        logger.error(f"Error in AI chat endpoint: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': f"An error occurred: {str(e)}"}), 500

@app.route('/api/ai/advanced_chat', methods=['POST'])
def advanced_ai_chat():
    """Advanced AI chat endpoint using LangChain SQL Agent for direct database querying"""
    try:
        # Check if LangChain is available
        if not LANGCHAIN_AVAILABLE:
            return jsonify({
                'success': False, 
                'error': "LangChain is not available. Please install required packages: pip install langchain langchain-openai langchain-community"
            }), 400
        
        # Check if OpenAI API key is valid
        if not openai_api_key:
            return jsonify({
                'success': False, 
                'error': "OpenAI API key is not configured correctly. Please set a valid API key."
            }), 400
            
        data = request.json
        if not data or 'message' not in data:
            return jsonify({'success': False, 'error': "Message field is required"}), 400
            
        user_message = data['message']
        logger.info(f"Processing advanced AI chat request: {user_message[:50]}...")
        
        # Create SQL agent
        try:
            agent_executor, db_langchain = create_sql_agent_instance()
        except Exception as e:
            logger.error(f"Failed to create SQL agent: {str(e)}")
            return jsonify({
                'success': False, 
                'error': f"Failed to initialize AI agent: {str(e)}"
            }), 500
        
        # Create enhanced prompt with business context
        enhanced_prompt = f"""
        You are an expert business intelligence analyst for a petrol station. You have direct access to the live database and can answer ANY question about the business data.
        
        {get_database_schema_info()}
        
        IMPORTANT SAFETY RULES:
        1. You can ONLY use SELECT statements - no INSERT, UPDATE, DELETE, DROP, etc.
        2. Always validate your SQL queries before execution
        3. If you need to calculate complex metrics, break them down into simple SELECT queries
        4. Always format currency amounts in Indian Rupees (Rs.)
        5. When showing dates, use a readable format (e.g., "August 3, 2025")
        6. If you can't find specific data, clearly state what's missing
        
        RESPONSE FORMAT:
        - Provide clear, actionable insights
        - Show actual numbers from the database
        - Include relevant context and trends
        - Use tables or bullet points for clarity
        
        User Question: {user_message}
        
        Please analyze the database and provide a comprehensive answer with real data.
        """
        
        try:
            # Execute the query using LangChain SQL agent
            logger.info("Executing LangChain SQL agent query...")
            result = agent_executor.invoke({"input": enhanced_prompt})
            
            # Extract the response
            ai_response = result.get('output', 'No response generated')
            
            # Log successful execution
            logger.info("LangChain SQL agent query executed successfully")
            
            # Save chat history
            chat_entry = ChatHistory(
                user_message=user_message,
                ai_response=ai_response
            )
            db.session.add(chat_entry)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'response': ai_response,
                'type': 'advanced_sql'
            })
            
        except Exception as e:
            logger.error(f"Error executing LangChain SQL agent: {str(e)}")
            
            # Fallback to regular AI chat if SQL agent fails
            logger.info("Falling back to regular AI chat due to SQL agent error")
            
            # Get business data summary as fallback
            business_data = get_business_data_summary()
            
            fallback_prompt = f"""
            You are a business intelligence analyst for a petrol station. Due to a technical issue with direct database access, 
            I'm providing you with a summary of the business data to answer the user's question.
            
            Business Data Summary:
            {json.dumps(business_data, indent=2)}
            
            User Question: {user_message}
            
            Please provide the best possible answer based on the available data summary. 
            If you need more specific data that's not in the summary, mention what additional information would be helpful.
            """
            
            try:
                # Use regular OpenAI API as fallback
                try:
                    # Try newer OpenAI client
                    client = openai.OpenAI(api_key=openai_api_key)
                    response = client.chat.completions.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": "You are a helpful business analyst."},
                            {"role": "user", "content": fallback_prompt}
                        ]
                    )
                except (ImportError, AttributeError):
                    # Fall back to legacy OpenAI client
                    response = openai.ChatCompletion.create(
                        model="gpt-4",
                        messages=[
                            {"role": "system", "content": "You are a helpful business analyst."},
                            {"role": "user", "content": fallback_prompt}
                        ]
                    )
                
                fallback_response = response.choices[0].message.content
                
                # Add note about fallback mode
                final_response = f"⚠️ **Note**: Advanced database querying is temporarily unavailable. This response is based on available data summary.\n\n{fallback_response}"
                
                # Save chat history
                chat_entry = ChatHistory(
                    user_message=user_message,
                    ai_response=final_response
                )
                db.session.add(chat_entry)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'response': final_response,
                    'type': 'fallback'
                })
                
            except Exception as fallback_error:
                logger.error(f"Fallback AI chat also failed: {str(fallback_error)}")
                return jsonify({
                    'success': False, 
                    'error': f"Both advanced and fallback AI systems failed: {str(e)}"
                }), 500
        
    except Exception as e:
        logger.error(f"Error in advanced AI chat endpoint: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': f"An error occurred: {str(e)}"}), 500

@app.route('/api/ai/insights', methods=['GET'])
def get_ai_insights():
    try:
        # Check if OpenAI API key is valid
        if not openai_api_key:
            logger.error("Invalid OpenAI API key. Please set a valid API key.")
            return jsonify({'success': False, 'error': "OpenAI API key is not configured correctly. Please set a valid API key."}), 400
        
        # Get a summary of business data
        business_data = get_business_data_summary()
        logger.debug(f"Business data summary retrieved: {len(str(business_data))} characters")

        if not business_data:
            return jsonify({
                'success': True,
                'insights': "No data available yet. Please add some daily sales or procurement entries to generate insights."
            })
        
        # Generate daily insights based on the summary
        system_prompt = f"""
You are an expert business intelligence analyst for a petroleum fuel station. Your primary goal is to provide accurate, transparent, and actionable insights based on the data provided. You must adhere to the following critical principles in every response:

**1. State Your Sources and Timeframes:**
Always begin your analysis by stating the date range of the data you are using. For example, "This analysis is based on data from August 2nd, 2025."

**2. Explicitly State the Number of Days for Averages:**
When calculating any average over time (like daily consumption or collections), you MUST explicitly state the exact number of days used in the calculation, which will be provided as `actual_days_of_data`.
* **Correct Example:** "Based on data from the last 3 days, the average daily HSD consumption is 1,500 Lts."
* **Incorrect Example:** "The average daily consumption is 1,500 Lts."

**3. Declare All Assumptions and Defaults:**
If you use any value that appears to be a hardcoded default or an assumption (e.g., tank capacities, fallback fuel rates), you MUST declare it.
* **Example for Tank Capacity:** "For the Tank Fill Optimization, I am using an *assumed* HSD1 tank capacity of 16,000 KL."
* **Example for Default Rates:** "The stock value was calculated using a *default* MS rate of Rs.108.56 as recent procurement data was not available."

**4. Report Data Conflicts with a "DATA ALERT":**
This is your most important task. If you detect a significant discrepancy between two data points (e.g., the fuel rate in a daily report is different from a hardcoded system rate), you MUST report this at the top of your response with a "🚨 DATA ALERT".
* **DATA ALERT Example:** "🚨 DATA ALERT: A potential data inconsistency was detected. The daily report for HSD shows a rate of Rs.96.46, but the system is using a default rate of Rs.92.85. The following calculations use the daily report rate. This discrepancy may impact financial accuracy."

**5. Show Your Work for Key Calculations:**
For critical metrics like payback timelines or inventory value, briefly show the formula you used to build trust and allow for verification.
* **Example:** "Credit Payback Timeline: 10 days (Calculation: Total Outstanding Rs.2,808,270 / Daily Collections Rs.310,593)."

By following these rules, you will act as a final quality check and ensure the user is never misled by potential issues in the underlying data.

Business Data Summary:
{json.dumps(business_data, indent=2)}
        """
        
        logger.info(f"Processing AI insights request")
        
        if USE_MOCK_RESPONSES:
            # Use mock response for insights
            logger.info("Using mock response for insights")
            
            # Create a comprehensive mock insights response
            insights_response = """# Daily Business Insights
            
## 1. STOCK DEPLETION FORECAST

| Fuel Type | Current Level | Daily Avg Usage | Days Remaining | Refill By |
|-----------|---------------|-----------------|----------------|-----------|
| MS        | 9,800 L       | 450 L/day       | 21.8 days      | Aug 25    |
| HSD       | 12,400 L      | 580 L/day       | 21.4 days      | Aug 24    |
| POWER     | 5,200 L       | 210 L/day       | 24.8 days      | Aug 28    |

Based on your 7-day consumption average, you have approximately 3 weeks of stock for all fuel types.

## 2. OPTIMAL REFILL SUGGESTION

🚚 **Recommended Order Schedule:**

| Fuel Type | Recommended Order Date | Optimal Order Quantity |
|-----------|------------------------|------------------------|
| MS        | August 15, 2025        | 8,000 liters          |
| HSD       | August 15, 2025        | 12,000 liters         |
| POWER     | August 20, 2025        | 6,000 liters          |

Considering HPCL's typical 2-day delivery timeframe, placing orders on these dates will maintain optimal inventory levels.

## 3. HPCL CREDIT PAYBACK TIMELINE

💰 **Current HPCL Credit Outstanding:** Rs.8,75,000

Based on your daily gross profit of approximately Rs.14,200:
- Estimated days to repay full amount: **62 days**
- Projected payback completion date: **October 4, 2025**

## 4. LIVE STOCK-TO-CASH CONVERSION

| Fuel Type | Current Stock | Current Rate | Cash Value |
|-----------|---------------|-------------|------------|
| MS        | 9,800 liters  | Rs.107.35/L   | Rs.10,52,030 |
| HSD       | 12,400 liters | Rs.92.85/L    | Rs.11,51,340 |
| POWER     | 5,200 liters  | Rs.112.45/L   | Rs.5,84,740  |
| **TOTAL** |               |             | **Rs.27,88,110** |

Your total fuel inventory is currently valued at approximately **Rs.27.88 lakhs**.
"""
            
            # Create a mock response object with just what we need
            class MockResponse:
                def __init__(self, content):
                    self.choices = [MockChoice(content)]
                    
            class MockChoice:
                def __init__(self, content):
                    self.message = MockMessage(content)
                    
            class MockMessage:
                def __init__(self, content):
                    self.content = content
                    
            response = MockResponse(insights_response)
            logger.info("Mock insights response generated successfully")
            
        else:
            # Try to use the real OpenAI API
            logger.info(f"Sending request to OpenAI API for insights with model: gpt-4")
            try:
                # Try newer OpenAI client first
                try:
                    logger.info("Initializing OpenAI client for insights")
                    client = openai.OpenAI(api_key=openai_api_key)
                    response = client.chat.completions.create(
                        model="gpt-4",  # Using GPT-4 for more advanced analysis
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": "Generate detailed business insights based on the provided data summary, covering all the required sections."}
                        ]
                    )
                except (ImportError, AttributeError) as e:
                    # Fall back to legacy client
                    logger.info(f"Using legacy OpenAI client for insights due to: {str(e)}")
                    response = openai.ChatCompletion.create(
                        model="gpt-4",  # Using GPT-4 for more advanced analysis
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": "Generate detailed business insights based on the provided data summary, covering all the required sections."}
                        ]
                    )
                logger.info(f"OpenAI API insights request successful")
            except Exception as api_error:
                logger.error(f"OpenAI API insights request failed with error: {str(api_error)}")
                raise api_error
        insights = response.choices[0].message.content

        return jsonify({
            'success': True,
            'insights': insights
        })
        
    except openai.APIError as e:
        logger.error(f"OpenAI API error (insights): {str(e)}")
        return jsonify({'success': False, 'error': f"OpenAI API error: {str(e)}. Please check your API key and try again."}), 400
    except openai.APIConnectionError as e:
        logger.error(f"OpenAI API connection error (insights): {str(e)}")
        return jsonify({'success': False, 'error': f"Failed to connect to OpenAI API: {str(e)}. Please check your internet connection."}), 400
    except openai.RateLimitError as e:
        logger.error(f"OpenAI API rate limit error (insights): {str(e)}")
        return jsonify({'success': False, 'error': f"OpenAI API rate limit exceeded: {str(e)}. Please try again later."}), 429
    except openai.AuthenticationError as e:
        logger.error(f"OpenAI API authentication error (insights): {str(e)}")
        return jsonify({'success': False, 'error': f"Authentication error with OpenAI API: {str(e)}. Please check your API key."}), 401
    except Exception as e:
        logger.error(f"Error in AI insights endpoint: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': f"An error occurred: {str(e)}"}), 400

# Dashboard Routes
@app.route('/api/dashboard/summary', methods=['GET'])
def dashboard_summary():
    try:
        start_date_str = request.args.get('start_date', date.today().isoformat())
        end_date_str = request.args.get('end_date', start_date_str)
        
        start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        logger.debug(f"Dashboard summary request - Start: {start_date_str}, End: {end_date_str}")
        
        # Query for entries within the date range
        daily_entries = DailyConsolidation.query.filter(
            DailyConsolidation.date >= start_date_obj,
            DailyConsolidation.date <= end_date_obj
        ).all()
        
        # Query for procurement entries within the date range
        procurement_entries = ProcurementData.query.filter(
            ProcurementData.invoice_date >= start_date_obj,
            ProcurementData.invoice_date <= end_date_obj
        ).all()
        
        # Calculate totals over the range
        total_fuel_sales = sum(entry.ms_amount + entry.hsd_amount + entry.power_amount for entry in daily_entries)
        total_collections = sum(entry.cash_collections + entry.card_collections + entry.paytm_collections + entry.hp_transactions for entry in daily_entries)
        variance = total_fuel_sales - total_collections

        # Get the HPCL credit outstanding amount specifically for the end date
        # This ensures we show the value for the exact date being viewed
        hpcl_credit_outstanding = 0.0
        hpcl_entry_date = end_date_obj
        hpcl_from_previous_date = False

        # 1. Find entries for the exact end date (which is what we're viewing)
        end_date_entries = DailyConsolidation.query.filter(
            DailyConsolidation.date == end_date_obj
        ).order_by(DailyConsolidation.created_at.desc()).all()

        if end_date_entries:
            # If entries exist for this date, use the latest one's value
            hpcl_credit_outstanding = end_date_entries[0].total_outstanding
            hpcl_entry_date = end_date_obj
            logger.debug(f"Found HPCL credit for end_date {end_date_obj}: {hpcl_credit_outstanding}")
        else:
            # 2. If no entry for end date, find the most recent entry BEFORE this date
            previous_entry = DailyConsolidation.query.filter(
                DailyConsolidation.date < end_date_obj
            ).order_by(DailyConsolidation.date.desc(), DailyConsolidation.created_at.desc()).first()
            
            if previous_entry:
                hpcl_credit_outstanding = previous_entry.total_outstanding
                hpcl_entry_date = previous_entry.date
                hpcl_from_previous_date = True
                logger.debug(f"Using previous entry from {previous_entry.date} with HPCL credit: {hpcl_credit_outstanding}")
            else:
                logger.debug(f"No entries found before {end_date_obj}, using default 0.0")
        
        # Calculate procurement totals
        total_procurement = sum(entry.total_amount for entry in procurement_entries)
        procurement_details = {
            'HSD': {'quantity': 0, 'amount': 0},
            'MS': {'quantity': 0, 'amount': 0},
            'POWER': {'quantity': 0, 'amount': 0}
        }
        for entry in procurement_entries:
            # Added robustness: handle potential casing issues or unexpected fuel types
            fuel_type_upper = entry.fuel_type.upper() if entry.fuel_type else None
            if fuel_type_upper in procurement_details:
                procurement_details[fuel_type_upper]['quantity'] += entry.quantity
                procurement_details[fuel_type_upper]['amount'] += entry.total_amount
        
        # Status checks are performed on the end_date of the range for UI clarity
        end_date_entries_for_status = DailyConsolidation.query.filter_by(date=end_date_obj).all()

        morning_shift_status = "Pending"
        if any(e.shift == 'Day' for e in end_date_entries_for_status):
            morning_shift_status = "Recorded"
            
        night_shift_status = "Pending"
        if any(e.shift == 'Night' for e in end_date_entries_for_status):
            night_shift_status = "Recorded"
        
        completion_status = get_daily_completion_status(end_date_obj)
        
        response_data = {
            'success': True,
            'data': {
                'start_date': start_date_str,
                'end_date': end_date_str,
                'total_fuel_sales': total_fuel_sales,
                'total_collections': total_collections,
                'variance': variance,
                'total_procurement': total_procurement,
                'procurement_details': procurement_details,
                'hpcl_credit_outstanding': hpcl_credit_outstanding,
                'hpcl_entry_date': hpcl_entry_date.isoformat(),
                'hpcl_from_previous_date': hpcl_from_previous_date,
                'completion_status': completion_status,
                'entries_count': len(end_date_entries_for_status),
                'morning_shift_status': morning_shift_status,
                'night_shift_status': night_shift_status
            }
        }
        
        logger.debug(f"Dashboard response: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error in dashboard summary: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400

# API endpoint for fetching HPCL credit specifically
@app.route('/api/business-day-info', methods=['GET'])
def get_business_day_info():
    """Get current business day information and suggested shift."""
    try:
        current_time = datetime.now()
        current_business_day = get_current_business_day()
        is_night_shift = is_night_shift_time()
        
        # Suggest the appropriate shift based on current time
        suggested_shift = 'Night' if is_night_shift else 'Day'
        
        # Get business day range
        start_time, end_time = get_business_day_range(current_business_day)
        
        return jsonify({
            'success': True,
            'data': {
                'current_business_day': current_business_day.isoformat(),
                'current_time': current_time.isoformat(),
                'suggested_shift': suggested_shift,
                'is_night_shift_time': is_night_shift,
                'business_day_start': start_time.isoformat(),
                'business_day_end': end_time.isoformat(),
                'shift_info': {
                    'day_shift': '8:30 AM - 8:30 PM',
                    'night_shift': '8:30 PM - 8:30 AM (next day)'
                }
            }
        })
    except Exception as e:
        logger.error(f"Error getting business day info: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard_data():
    try:
        # Get date parameters from request (if provided) - this should be a business date
        target_date_str = request.args.get('date')
        
        if target_date_str:
            try:
                target_business_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            except ValueError:
                target_business_date = get_current_business_day()
        else:
            # Use current business day if no date specified
            target_business_date = get_current_business_day()
        
        logger.debug(f"Fetching dashboard data for business date: {target_business_date}")
        
        # Find entries for the specific business date (both day and night shifts)
        date_entries = DailyConsolidation.query.filter(
            DailyConsolidation.date == target_business_date
        ).order_by(DailyConsolidation.created_at.desc()).all()
        
        hpcl_credit_outstanding = 0.0
        entry_date_str = target_business_date.isoformat()
        from_previous_date = False
        
        if date_entries:
            # If entries exist for this business date, use the latest one's outstanding value
            latest_entry = date_entries[0]
            hpcl_credit_outstanding = float(latest_entry.total_outstanding)
            logger.debug(f"Found HPCL credit outstanding for business date {target_business_date}: {hpcl_credit_outstanding}")
        else:
            # If no entry for this business date, find the most recent entry before this date
            previous_entry = DailyConsolidation.query.filter(
                DailyConsolidation.date < target_business_date
            ).order_by(DailyConsolidation.date.desc(), DailyConsolidation.created_at.desc()).first()
            
            if previous_entry:
                hpcl_credit_outstanding = float(previous_entry.total_outstanding)
                entry_date_str = previous_entry.date.isoformat()
                from_previous_date = True
                logger.debug(f"Using previous entry from business date {previous_entry.date} with HPCL credit: {hpcl_credit_outstanding}")
            else:
                logger.debug(f"No entries found before business date {target_business_date}, using default 0.0")
        
        return jsonify({
            'status': 'success',
            'dashboardData': {
                'hpcl_credit_outstanding': hpcl_credit_outstanding,
                'view_business_date': target_business_date.isoformat(),
                'entry_date': entry_date_str,
                'from_previous_date': from_previous_date,
                'is_business_day_logic': True
            }
        })
    except Exception as e:
        logger.error(f"Error fetching dashboard data: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Frontend serving routes
@app.route('/')
def serve_frontend():
    return render_template('index.html')

@app.route('/<path:path>')
def serve_frontend_routes(path):
    # For React Router - serve index.html for all frontend routes
    if path.startswith('api/'):
        return jsonify({'error': 'API endpoint not found'}), 404
    return render_template('index.html')

@app.route('/assets/<path:filename>')
def serve_assets(filename):
    return send_from_directory('static/assets', filename)

# ===============================
# GOOGLE SHEETS INTEGRATION APIs
# ===============================

@app.route('/api/google-sheets/status', methods=['GET'])
def google_sheets_status():
    """Get Google Sheets integration status"""
    try:
        if not GOOGLE_SHEETS_AVAILABLE:
            return jsonify({
                'available': False,
                'error': 'Google Sheets integration not installed'
            })
        
        status = {
            'available': GOOGLE_SHEETS_AVAILABLE,
            'sync_initialized': google_sync is not None,
            'ai_chat_initialized': google_ai_chat is not None,
            'owner_email': 'hemanth.gajjala88@gmail.com'
        }
        
        if google_sync:
            try:
                spreadsheet_info = google_sync.get_spreadsheet_info()
                status.update({
                    'spreadsheet_url': spreadsheet_info.get('url'),
                    'spreadsheet_title': spreadsheet_info.get('title'),
                    'last_sync': spreadsheet_info.get('last_sync')
                })
            except Exception as e:
                status['sync_error'] = str(e)
        
        return jsonify(status)
        
    except Exception as e:
        logger.error(f"Error getting Google Sheets status: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/google-sheets/sync', methods=['POST'])
def sync_to_google_sheets():
    """Manually sync database to Google Sheets"""
    try:
        if not google_sync:
            return jsonify({'error': 'Google Sheets sync not initialized'}), 400
        
        # Perform sync
        success = google_sync.sync_all_tables()
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Data synced to Google Sheets successfully',
                'timestamp': datetime.now().isoformat(),
                'spreadsheet_url': google_sync.get_spreadsheet_info().get('url')
            })
        else:
            return jsonify({'error': 'Sync failed'}), 500
            
    except Exception as e:
        logger.error(f"Error syncing to Google Sheets: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/google-sheets/ai-chat', methods=['POST'])
def google_sheets_ai_chat():
    """AI Chat using Google Sheets data"""
    try:
        if not google_ai_chat:
            return jsonify({'error': 'Google Sheets AI Chat not initialized'}), 400
        
        data = request.get_json()
        user_message = data.get('message', '')
        
        if not user_message:
            return jsonify({'error': 'Message is required'}), 400
        
        # Get AI response using Google Sheets data
        ai_response = google_ai_chat.chat_with_sheets(user_message)
        
        # Save to chat history (optional)
        try:
            chat_entry = ChatHistory(
                user_message=user_message,
                ai_response=ai_response,
                timestamp=datetime.now()
            )
            db.session.add(chat_entry)
            db.session.commit()
        except Exception as e:
            logger.warning(f"Failed to save chat history: {str(e)}")
        
        return jsonify({
            'response': ai_response,
            'timestamp': datetime.now().isoformat(),
            'data_source': 'Google Sheets',
            'spreadsheet_url': google_ai_chat.data_provider.sync_client.get_sheet_url()
        })
        
    except Exception as e:
        logger.error(f"Google Sheets AI Chat error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/google-sheets/data/<table_name>', methods=['GET'])
def get_google_sheets_data(table_name):
    """Get data from Google Sheets for a specific table"""
    try:
        if not google_sync:
            return jsonify({'error': 'Google Sheets sync not initialized'}), 400
        
        # Map table names to sheet names
        table_mapping = {
            'daily-consolidation': 'Daily Consolidation',
            'procurement': 'Procurement Data',
            'tank-readings': 'Tank Readings',
            'customer-credit': 'Customer Credit',
            'hpcl-payments': 'HPCL Payments',
            'chat-history': 'Chat History'
        }
        
        sheet_name = table_mapping.get(table_name)
        if not sheet_name:
            return jsonify({'error': f'Unknown table: {table_name}'}), 400
        
        # Get data from Google Sheets
        data = google_sync.get_sheet_data(sheet_name)
        
        return jsonify({
            'table_name': table_name,
            'sheet_name': sheet_name,
            'data': data,
            'count': len(data),
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error getting Google Sheets data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/google-sheets/setup', methods=['POST'])
def setup_google_sheets():
    """Setup Google Sheets integration"""
    try:
        global google_sync, google_ai_chat
        
        if not GOOGLE_SHEETS_AVAILABLE:
            return jsonify({
                'error': 'Google Sheets integration not available',
                'instructions': 'Install required packages: pip install -r requirements_google.txt'
            }), 400
        
        # Initialize Google Drive sync
        google_sync = GoogleDriveDBSync(
            owner_email="hemanth.gajjala88@gmail.com",
            spreadsheet_name="Petrol Station Data - Live"
        )
        
        # Setup or access spreadsheet
        if google_sync.setup_or_access_spreadsheet():
            # Initial sync
            sync_success = google_sync.sync_all_tables()
            
            # Initialize AI chat
            if openai_api_key:
                google_ai_chat = GoogleSheetsAIChat(
                    openai_api_key=openai_api_key,
                    owner_email="hemanth.gajjala88@gmail.com"
                )
            
            return jsonify({
                'success': True,
                'message': 'Google Sheets integration setup successfully',
                'spreadsheet_url': google_sync.get_sheet_url(),
                'sync_success': sync_success,
                'ai_chat_available': google_ai_chat is not None
            })
        else:
            return jsonify({'error': 'Failed to setup Google Sheets'}), 500
            
    except Exception as e:
        logger.error(f"Error setting up Google Sheets: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Export data endpoints
@app.route('/api/export/<data_type>', methods=['GET'])
def export_data(data_type):
    """Export data as CSV"""
    try:
        import csv
        import io
        from flask import make_response
        
        if data_type == 'daily-consolidation':
            entries = DailyConsolidation.query.order_by(DailyConsolidation.date.desc()).all()
            fieldnames = [
                'id', 'date', 'shift', 'manager', 'ms_rate', 'ms_quantity', 'ms_amount',
                'hsd_rate', 'hsd_quantity', 'hsd_amount', 'power_rate', 'power_quantity', 'power_amount',
                'cash_collections', 'card_collections', 'paytm_collections', 'hp_transactions',
                'hpcl_payment', 'created_at'
            ]
            filename = f'daily_consolidation_{datetime.now().strftime("%Y%m%d")}.csv'
            
        elif data_type == 'procurement':
            entries = Procurement.query.order_by(Procurement.date.desc()).all()
            fieldnames = [
                'id', 'date', 'product', 'quantity', 'rate', 'amount', 'supplier', 'invoice_number', 'created_at'
            ]
            filename = f'procurement_entries_{datetime.now().strftime("%Y%m%d")}.csv'
            
        elif data_type == 'tank-readings':
            entries = TankReading.query.order_by(TankReading.date.desc()).all()
            fieldnames = [
                'id', 'date', 'hsd1_tank', 'hsd2_tank', 'ms1_tank', 'ms2_tank', 'power1_tank', 'created_at'
            ]
            filename = f'tank_readings_{datetime.now().strftime("%Y%m%d")}.csv'
            
        elif data_type == 'customer-credit':
            entries = CustomerCredit.query.order_by(CustomerCredit.date.desc()).all()
            fieldnames = [
                'id', 'customer_name', 'date', 'transaction_type', 'fuel_type', 'quantity', 
                'rate', 'amount', 'balance', 'notes', 'created_at'
            ]
            filename = f'customer_credit_{datetime.now().strftime("%Y%m%d")}.csv'
            
        else:
            return jsonify({'error': 'Invalid data type'}), 400
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for entry in entries:
            row = {}
            for field in fieldnames:
                value = getattr(entry, field, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    value = value.strftime('%Y-%m-%d')
                row[field] = value
            writer.writerow(row)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting {data_type}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/all', methods=['GET'])
def export_all_data():
    """Export all data as a ZIP file containing multiple CSV files"""
    try:
        import zipfile
        import io
        import csv
        from flask import make_response
        
        # Create a BytesIO object to hold the zip file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Export daily consolidation
            entries = DailyConsolidation.query.order_by(DailyConsolidation.date.desc()).all()
            if entries:
                output = io.StringIO()
                fieldnames = [
                    'id', 'date', 'shift', 'manager', 'ms_rate', 'ms_quantity', 'ms_amount',
                    'hsd_rate', 'hsd_quantity', 'hsd_amount', 'power_rate', 'power_quantity', 'power_amount',
                    'cash_collections', 'card_collections', 'paytm_collections', 'hp_transactions',
                    'hpcl_payment', 'created_at'
                ]
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for entry in entries:
                    row = {}
                    for field in fieldnames:
                        value = getattr(entry, field, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, date):
                            value = value.strftime('%Y-%m-%d')
                        row[field] = value
                    writer.writerow(row)
                zip_file.writestr('daily_consolidation.csv', output.getvalue())
            
            # Export procurement
            entries = Procurement.query.order_by(Procurement.date.desc()).all()
            if entries:
                output = io.StringIO()
                fieldnames = [
                    'id', 'date', 'product', 'quantity', 'rate', 'amount', 'supplier', 'invoice_number', 'created_at'
                ]
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for entry in entries:
                    row = {}
                    for field in fieldnames:
                        value = getattr(entry, field, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, date):
                            value = value.strftime('%Y-%m-%d')
                        row[field] = value
                    writer.writerow(row)
                zip_file.writestr('procurement_entries.csv', output.getvalue())
            
            # Export tank readings
            entries = TankReading.query.order_by(TankReading.date.desc()).all()
            if entries:
                output = io.StringIO()
                fieldnames = [
                    'id', 'date', 'hsd1_tank', 'hsd2_tank', 'ms1_tank', 'ms2_tank', 'power1_tank', 'created_at'
                ]
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for entry in entries:
                    row = {}
                    for field in fieldnames:
                        value = getattr(entry, field, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, date):
                            value = value.strftime('%Y-%m-%d')
                        row[field] = value
                    writer.writerow(row)
                zip_file.writestr('tank_readings.csv', output.getvalue())
            
            # Export customer credit
            entries = CustomerCredit.query.order_by(CustomerCredit.date.desc()).all()
            if entries:
                output = io.StringIO()
                fieldnames = [
                    'id', 'customer_name', 'date', 'transaction_type', 'fuel_type', 'quantity', 
                    'rate', 'amount', 'balance', 'notes', 'created_at'
                ]
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for entry in entries:
                    row = {}
                    for field in fieldnames:
                        value = getattr(entry, field, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, date):
                            value = value.strftime('%Y-%m-%d')
                        row[field] = value
                    writer.writerow(row)
                zip_file.writestr('customer_credit.csv', output.getvalue())
        
        zip_buffer.seek(0)
        
        # Create response
        response = make_response(zip_buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=petrol_station_data_{datetime.now().strftime("%Y%m%d")}.zip'
        response.headers['Content-Type'] = 'application/zip'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting all data: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Business Intelligence Endpoints
@app.route('/api/analytics/current-stock', methods=['GET'])
def get_current_stock():
    """Get current stock levels from latest daily consolidation entry."""
    try:
        latest_entry = DailyConsolidation.query.order_by(DailyConsolidation.id.desc()).first()
        
        if not latest_entry:
            return jsonify({
                'ms1_tank': 8500,
                'ms2_tank': 0,
                'hsd1_tank': 12000,
                'hsd2_tank': 0,
                'power1_tank': 3500
            })
        
        return jsonify({
            'ms1_tank': latest_entry.ms1_tank or 0,
            'ms2_tank': latest_entry.ms2_tank or 0,
            'hsd1_tank': latest_entry.hsd1_tank or 0,
            'hsd2_tank': latest_entry.hsd2_tank or 0,
            'power1_tank': latest_entry.power1_tank or 0
        })
        
    except Exception as e:
        logger.error(f"Error getting current stock: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/sales-trends', methods=['GET'])
def get_sales_trends():
    """Get sales trends and patterns with actual rates from daily entries."""
    try:
        # Get last 30 days of data
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
        
        entries = DailyConsolidation.query.filter(
            DailyConsolidation.date >= start_date,
            DailyConsolidation.date <= end_date
        ).order_by(DailyConsolidation.date.asc()).all()
        
        sales_data = []
        for entry in entries:
            total_sales = (entry.ms_amount or 0) + (entry.hsd_amount or 0) + (entry.power_amount or 0)
            sales_data.append({
                'date': entry.date.isoformat(),
                'total_sales': total_sales,
                'ms_amount': entry.ms_amount or 0,
                'hsd_amount': entry.hsd_amount or 0,
                'power_amount': entry.power_amount or 0,
                'ms_quantity': entry.ms_quantity or 0,
                'hsd_quantity': entry.hsd_quantity or 0,
                'power_quantity': entry.power_quantity or 0,
                'ms_rate': entry.ms_rate or 106.50,
                'hsd_rate': entry.hsd_rate or 94.20,
                'power_rate': entry.power_rate or 89.30,
                'manager': getattr(entry, 'manager', 'Unknown'),
                'shift': getattr(entry, 'shift', 'Day'),
                'cash_collections': entry.cash_collections or 0,
                'card_collections': entry.card_collections or 0,
                'paytm_collections': entry.paytm_collections or 0,
                'hpcl_payment': entry.hpcl_payment or 0
            })
        
        return jsonify(sales_data)
        
    except Exception as e:
        logger.error(f"Error getting sales trends: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/business-intelligence/credit-timeline', methods=['GET'])
def calculate_credit_timeline():
    """Calculate HPCL credit payback timeline using actual data from daily entries."""
    try:
        # Get the latest HPCL outstanding credit from the most recent daily entry
        latest_entry = DailyConsolidation.query.order_by(
            DailyConsolidation.date.desc(), 
            DailyConsolidation.id.desc()
        ).first()
        
        if latest_entry and hasattr(latest_entry, 'total_outstanding') and latest_entry.total_outstanding:
            # Use the actual outstanding credit from daily entry
            outstanding_credit = float(latest_entry.total_outstanding)
            logger.info(f"Using HPCL outstanding credit from latest daily entry: ₹{outstanding_credit:,.0f}")
        else:
            # Fallback: Calculate from procurement and payments if daily entry doesn't have outstanding
            procurement_entries = ProcurementData.query.all()
            total_procurement_cost = sum(entry.quantity * entry.rate for entry in procurement_entries)
            
            # Get all HPCL payments from daily consolidation
            daily_entries = DailyConsolidation.query.all()
            total_hpcl_payments = sum(entry.hpcl_payment or 0 for entry in daily_entries)
            
            outstanding_credit = total_procurement_cost - total_hpcl_payments
            logger.info(f"Calculated HPCL outstanding credit from procurement/payments: ₹{outstanding_credit:,.0f}")
        
        # Calculate average daily cash flow (last 30 days)
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
        
        recent_entries = DailyConsolidation.query.filter(
            DailyConsolidation.date >= start_date,
            DailyConsolidation.date <= end_date
        ).all()
        
        if recent_entries:
            total_cash_flow = sum(
                (entry.cash_collections or 0) + 
                (entry.card_collections or 0) + 
                (entry.paytm_collections or 0) 
                for entry in recent_entries
            )
            avg_daily_cash_flow = total_cash_flow / len(recent_entries)
        else:
            avg_daily_cash_flow = 180000  # Default
        
        # Business parameters
        CREDIT_PAYMENT_RATIO = 0.3  # 30% of daily cash flow for credit payment
        daily_credit_payment_capacity = avg_daily_cash_flow * CREDIT_PAYMENT_RATIO
        
        if daily_credit_payment_capacity > 0:
            payback_days = max(0, int(outstanding_credit / daily_credit_payment_capacity))
        else:
            payback_days = 999
        
        payback_date = (datetime.now() + timedelta(days=payback_days)).strftime('%b %d, %Y')
        
        return jsonify({
            'outstandingCredit': outstanding_credit,
            'dailyCashFlow': avg_daily_cash_flow,
            'dailyCreditPaymentCapacity': daily_credit_payment_capacity,
            'paybackDays': payback_days,
            'paybackDate': payback_date
        })
        
    except Exception as e:
        logger.error(f"Error calculating credit timeline: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/business-intelligence/sales-trends', methods=['GET'])
def calculate_sales_trends():
    """Calculate comprehensive sales trend analysis with multi-timeframe analysis and volatility measurement."""
    try:
        import numpy as np
        from datetime import timedelta
        import statistics
        
        # Configuration
        VOLATILITY_THRESHOLDS = {
            'low': 15,      # CV < 15% = Low volatility
            'medium': 30,   # CV 15-30% = Medium volatility  
            'high': 50,     # CV 30-50% = High volatility
            'extreme': 100  # CV > 50% = Extreme volatility
        }
        
        # Get all daily entries
        daily_entries = DailyConsolidation.query.order_by(DailyConsolidation.date.asc()).all()
        
        if not daily_entries:
            return jsonify({'error': 'No sales data available'}), 404
        
        # Prepare data for analysis
        sales_data = []
        for entry in daily_entries:
            ms_amount = float(entry.ms_amount or 0)
            hsd_amount = float(entry.hsd_amount or 0)
            power_amount = float(entry.power_amount or 0)
            
            sales_data.append({
                'date': entry.date,
                'ms_amount': ms_amount,
                'hsd_amount': hsd_amount,
                'power_amount': power_amount,
                'total_amount': ms_amount + hsd_amount + power_amount
            })
        
        def calculate_coefficient_of_variation(values):
            """Calculate coefficient of variation as percentage."""
            if not values or len(values) < 2:
                return 0.0
            mean_val = statistics.mean(values)
            if mean_val == 0:
                return 0.0
            std_val = statistics.stdev(values)
            return (std_val / mean_val) * 100
        
        def classify_volatility(cv):
            """Classify volatility based on coefficient of variation."""
            if cv < VOLATILITY_THRESHOLDS['low']:
                return 'Low'
            elif cv < VOLATILITY_THRESHOLDS['medium']:
                return 'Medium'
            elif cv < VOLATILITY_THRESHOLDS['high']:
                return 'High'
            else:
                return 'Extreme'
        
        def calculate_trend_slope(values):
            """Calculate linear regression slope for trend analysis."""
            if len(values) < 2:
                return 0.0
            
            n = len(values)
            x = list(range(n))
            y = values
            
            # Linear regression slope calculation
            sum_x = sum(x)
            sum_y = sum(y)
            sum_xy = sum(x[i] * y[i] for i in range(n))
            sum_x_squared = sum(x[i] ** 2 for i in range(n))
            
            denominator = n * sum_x_squared - sum_x ** 2
            if denominator == 0:
                return 0.0
            
            slope = (n * sum_xy - sum_x * sum_y) / denominator
            return slope
        
        def classify_trend(slope, mean_value):
            """Classify trend based on slope and mean value."""
            if mean_value == 0:
                return 'Stable'
            
            # Convert slope to percentage change per day
            percent_change_per_day = (slope / mean_value) * 100
            
            if percent_change_per_day < -2:
                return 'Strong Decline'
            elif percent_change_per_day < -0.5:
                return 'Moderate Decline'
            elif percent_change_per_day > 2:
                return 'Strong Growth'
            elif percent_change_per_day > 0.5:
                return 'Moderate Growth'
            else:
                return 'Stable'
        
        def calculate_period_over_period_change(recent_values, previous_values):
            """Calculate percentage change between periods."""
            if not recent_values or not previous_values:
                return 0.0
            
            recent_avg = statistics.mean(recent_values)
            previous_avg = statistics.mean(previous_values)
            
            if previous_avg == 0:
                return 0.0
            
            return ((recent_avg - previous_avg) / previous_avg) * 100
        
        # Daily Analysis
        daily_analysis = {}
        fuel_types = ['ms_amount', 'hsd_amount', 'power_amount', 'total_amount']
        fuel_names = ['Motor Spirit', 'High Speed Diesel', 'Power/Kerosene', 'Total Sales']
        
        for i, fuel_type in enumerate(fuel_types):
            values = [entry[fuel_type] for entry in sales_data if entry[fuel_type] > 0]
            
            if values:
                mean_val = statistics.mean(values)
                median_val = statistics.median(values)
                std_val = statistics.stdev(values) if len(values) > 1 else 0
                cv = calculate_coefficient_of_variation(values)
                slope = calculate_trend_slope(values)
                trend_class = classify_trend(slope, mean_val)
                volatility_class = classify_volatility(cv)
                
                # Recent vs Previous comparison (last 7 vs previous 7 days)
                recent_values = values[-7:] if len(values) >= 7 else values
                previous_values = values[-14:-7] if len(values) >= 14 else values[:-7] if len(values) > 7 else []
                
                period_change = calculate_period_over_period_change(recent_values, previous_values)
                
                daily_analysis[fuel_type] = {
                    'fuel_name': fuel_names[i],
                    'mean_daily_sales': round(mean_val, 2),
                    'median_daily_sales': round(median_val, 2),
                    'std_daily_sales': round(std_val, 2),
                    'coefficient_of_variation': round(cv, 1),
                    'volatility_classification': volatility_class,
                    'trend_slope': round(slope, 2),
                    'trend_classification': trend_class,
                    'recent_vs_previous_change': round(period_change, 1),
                    'data_points': len(values)
                }
        
        # Weekly Analysis
        weekly_data = {}
        if len(sales_data) >= 7:
            # Group by weeks
            weeks = []
            current_week = []
            
            for i, entry in enumerate(sales_data):
                current_week.append(entry)
                if len(current_week) == 7:
                    weeks.append(current_week)
                    current_week = []
            
            if weeks:
                for fuel_type in fuel_types:
                    weekly_totals = []
                    for week in weeks:
                        weekly_total = sum(day[fuel_type] for day in week)
                        weekly_totals.append(weekly_total)
                    
                    if weekly_totals:
                        weekly_cv = calculate_coefficient_of_variation(weekly_totals)
                        weekly_mean = statistics.mean(weekly_totals)
                        
                        # Week-over-week changes
                        wow_changes = []
                        for i in range(1, len(weekly_totals)):
                            if weekly_totals[i-1] > 0:
                                change = ((weekly_totals[i] - weekly_totals[i-1]) / weekly_totals[i-1]) * 100
                                wow_changes.append(change)
                        
                        positive_weeks = sum(1 for change in wow_changes if change > 0)
                        negative_weeks = sum(1 for change in wow_changes if change < 0)
                        
                        weekly_data[fuel_type] = {
                            'mean_weekly_sales': round(weekly_mean, 2),
                            'weekly_volatility': round(weekly_cv, 1),
                            'mean_wow_change': round(statistics.mean(wow_changes), 1) if wow_changes else 0,
                            'positive_weeks': positive_weeks,
                            'negative_weeks': negative_weeks,
                            'total_weeks': len(weekly_totals)
                        }
        
        # Volatility Ranking
        volatility_ranking = []
        for fuel_type in fuel_types:
            if fuel_type in daily_analysis:
                volatility_ranking.append({
                    'fuel_type': daily_analysis[fuel_type]['fuel_name'],
                    'fuel_code': fuel_type,
                    'volatility_score': daily_analysis[fuel_type]['coefficient_of_variation'],
                    'volatility_classification': daily_analysis[fuel_type]['volatility_classification'],
                    'trend_classification': daily_analysis[fuel_type]['trend_classification']
                })
        
        # Sort by volatility score (highest first)
        volatility_ranking.sort(key=lambda x: x['volatility_score'], reverse=True)
        
        # Executive Summary
        total_data = daily_analysis.get('total_amount', {})
        overall_trend = total_data.get('trend_classification', 'Stable')
        most_volatile = volatility_ranking[0] if volatility_ranking else None
        least_volatile = volatility_ranking[-1] if volatility_ranking else None
        avg_daily_sales = total_data.get('mean_daily_sales', 0)
        
        executive_summary = {
            'overall_trend_daily': overall_trend,
            'most_volatile_fuel': most_volatile['fuel_type'] if most_volatile else 'N/A',
            'most_volatile_cv': most_volatile['volatility_score'] if most_volatile else 0,
            'least_volatile_fuel': least_volatile['fuel_type'] if least_volatile else 'N/A',
            'least_volatile_cv': least_volatile['volatility_score'] if least_volatile else 0,
            'average_daily_sales': round(avg_daily_sales, 0),
            'total_data_points': len(sales_data)
        }
        
        # Return comprehensive analysis
        return jsonify({
            'timestamp': datetime.now().isoformat(),
            'executive_summary': executive_summary,
            'daily_analysis': daily_analysis,
            'weekly_analysis': weekly_data,
            'volatility_ranking': volatility_ranking,
            'configuration': {
                'volatility_thresholds': VOLATILITY_THRESHOLDS,
                'analysis_period_days': len(sales_data)
            }
        })
        
    except Exception as e:
        logger.error(f"Error calculating sales trends: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/business-intelligence/stock-depletion', methods=['GET'])
def enhanced_stock_depletion_forecast():
    """Enhanced stock depletion forecast with multi-scenario analysis."""
    try:
        import numpy as np
        from datetime import datetime, timedelta
        
        # Configuration parameters
        TANK_CAPACITIES = {'HSD': 12000, 'MS': 15000, 'Power': 8000}
        MINIMUM_THRESHOLD = 0.15
        REORDER_THRESHOLD = 0.20
        SAFETY_BUFFER_DAYS = 2
        LEAD_TIME_DAYS = 3
        ANALYSIS_LOOKBACK_DAYS = 14
        
        # Get current tank levels
        latest_entry = DailyConsolidation.query.order_by(
            DailyConsolidation.date.desc(), 
            DailyConsolidation.id.desc()
        ).first()
        
        if not latest_entry:
            return jsonify({'error': 'No tank data available'}), 404
        
        # Calculate current stock levels
        current_stock = {
            'HSD': {
                'tank1': float(latest_entry.hsd1_tank or 0),
                'tank2': float(latest_entry.hsd2_tank or 0),
                'total_liters': float((latest_entry.hsd1_tank or 0) + (latest_entry.hsd2_tank or 0)),
                'capacity': TANK_CAPACITIES['HSD']
            },
            'MS': {
                'tank1': float(latest_entry.ms1_tank or 0),
                'tank2': float(latest_entry.ms2_tank or 0),
                'total_liters': float((latest_entry.ms1_tank or 0) + (latest_entry.ms2_tank or 0)),
                'capacity': TANK_CAPACITIES['MS']
            },
            'Power': {
                'tank1': float(latest_entry.power1_tank or 0),
                'total_liters': float(latest_entry.power1_tank or 0),
                'capacity': TANK_CAPACITIES['Power']
            }
        }
        
        # Calculate percentages
        for fuel_type in current_stock:
            current_stock[fuel_type]['percentage'] = (
                current_stock[fuel_type]['total_liters'] / current_stock[fuel_type]['capacity']
            ) * 100
        
        # Get historical consumption data (last 14 days)
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=ANALYSIS_LOOKBACK_DAYS)
        
        historical_entries = DailyConsolidation.query.filter(
            DailyConsolidation.date >= start_date,
            DailyConsolidation.date <= end_date
        ).order_by(DailyConsolidation.date.asc()).all()
        
        if len(historical_entries) < 3:
            # Fallback to demo data if insufficient historical data
            consumption_data = {
                'HSD': [2200, 2400, 2150, 2600, 2300, 2100, 2350, 2450, 2200, 2500, 2250, 2380, 2320, 2480],
                'MS': [1800, 1950, 1750, 2100, 1850, 1700, 1900, 1980, 1800, 2000, 1820, 1940, 1880, 1960],
                'Power': [350, 380, 320, 420, 360, 300, 380, 390, 350, 400, 340, 380, 370, 390]
            }
        else:
            # Calculate actual consumption from historical data
            consumption_data = {'HSD': [], 'MS': [], 'Power': []}
            
            for entry in historical_entries:
                consumption_data['HSD'].append(float(entry.hsd_quantity or 0))
                consumption_data['MS'].append(float(entry.ms_quantity or 0))
                consumption_data['Power'].append(float(entry.power_quantity or 0))
        
        forecast_results = {}
        
        for fuel_type in ['HSD', 'MS', 'Power']:
            consumption_list = consumption_data[fuel_type]
            
            if not consumption_list or all(x == 0 for x in consumption_list):
                consumption_list = [2000, 2100, 1950, 2200, 2050] if fuel_type == 'HSD' else [1500, 1600, 1450, 1650, 1550] if fuel_type == 'MS' else [300, 320, 280, 340, 310]
            
            # Statistical analysis
            consumption_array = np.array(consumption_list)
            daily_average = float(np.mean(consumption_array))
            daily_median = float(np.median(consumption_array))
            std_deviation = float(np.std(consumption_array))
            daily_max = float(np.max(consumption_array))
            percentile_75 = float(np.percentile(consumption_array, 75))
            
            # Conservative consumption rate (75th percentile or mean + 0.5*std)
            conservative_daily = max(percentile_75, daily_average + 0.5 * std_deviation)
            
            # Trend analysis (linear regression slope)
            if len(consumption_array) > 1:
                x = np.arange(len(consumption_array))
                slope = float(np.polyfit(x, consumption_array, 1)[0])
                trend_adjusted_rate = daily_average + (slope * 7)  # 7-day projection
            else:
                slope = 0
                trend_adjusted_rate = daily_average
            
            # Volatility calculation
            volatility = (std_deviation / daily_average) * 100 if daily_average > 0 else 0
            
            # Calculate depletion scenarios
            current_liters = current_stock[fuel_type]['total_liters']
            capacity = current_stock[fuel_type]['capacity']
            reorder_threshold_liters = capacity * REORDER_THRESHOLD
            minimum_threshold_liters = capacity * MINIMUM_THRESHOLD
            
            # Scenario calculations
            scenarios = {}
            
            # Scenario 1: Average Consumption
            if daily_average > 0:
                scenarios['average'] = max(0, (current_liters - reorder_threshold_liters) / daily_average)
            else:
                scenarios['average'] = 999
            
            # Scenario 2: Conservative Consumption
            if conservative_daily > 0:
                scenarios['conservative'] = max(0, (current_liters - reorder_threshold_liters) / conservative_daily)
            else:
                scenarios['conservative'] = 999
            
            # Scenario 3: Peak Consumption
            if daily_max > 0:
                scenarios['peak'] = max(0, (current_liters - reorder_threshold_liters) / daily_max)
            else:
                scenarios['peak'] = 999
            
            # Scenario 4: Trend-Adjusted
            if trend_adjusted_rate > 0:
                scenarios['trend_adjusted'] = max(0, (current_liters - reorder_threshold_liters) / trend_adjusted_rate)
            else:
                scenarios['trend_adjusted'] = 999
            
            # Days to minimum threshold
            days_to_minimum = max(0, (current_liters - minimum_threshold_liters) / conservative_daily) if conservative_daily > 0 else 999
            
            # Status determination
            current_percentage = current_stock[fuel_type]['percentage']
            days_to_reorder = scenarios['conservative']
            
            if current_percentage <= 15:
                status_level = "CRITICAL"
                status_color = "🔴"
                urgency = "HIGH"
                message = "Below minimum threshold"
            elif current_percentage <= 20:
                status_level = "URGENT"
                status_color = "🟠"
                urgency = "HIGH"
                message = "At reorder threshold"
            elif days_to_reorder <= (LEAD_TIME_DAYS + SAFETY_BUFFER_DAYS):
                status_level = "WARNING"
                status_color = "🟡"
                urgency = "MEDIUM"
                message = f"Will reach reorder threshold in {days_to_reorder:.1f} days"
            elif days_to_reorder <= 10:
                status_level = "CAUTION"
                status_color = "🟡"
                urgency = "LOW"
                message = "Monitor closely"
            else:
                status_level = "GOOD"
                status_color = "🟢"
                urgency = "LOW"
                message = "Stock levels adequate"
            
            # Recommendations
            immediate_actions = []
            if status_level in ["CRITICAL", "URGENT"]:
                immediate_actions.append(f"Order {fuel_type} fuel immediately")
                immediate_actions.append("Confirm supplier availability and delivery schedule")
            elif status_level == "WARNING":
                immediate_actions.append(f"Prepare {fuel_type} fuel order")
                immediate_actions.append("Schedule delivery within 2-3 days")
            elif status_level == "CAUTION":
                immediate_actions.append(f"Monitor {fuel_type} consumption closely")
                immediate_actions.append("Plan fuel order for next 3-5 days")
            
            # Order recommendations
            optimal_quantity = max(0, capacity * 0.8 - current_liters)  # Fill to 80%
            minimum_quantity = max(0, reorder_threshold_liters * 2 - current_liters)
            
            # Monitoring frequency based on volatility
            if volatility > 30:
                monitoring_frequency = "Daily monitoring required"
            elif volatility > 15:
                monitoring_frequency = "Monitor every 2-3 days"
            else:
                monitoring_frequency = "Weekly monitoring sufficient"
            
            forecast_results[fuel_type] = {
                'current_stock': {
                    'total_liters': current_liters,
                    'percentage': current_percentage,
                    'capacity': capacity,
                    'above_minimum': current_percentage > 15,
                    'above_reorder': current_percentage > 20
                },
                'consumption_patterns': {
                    'daily_average': daily_average,
                    'daily_median': daily_median,
                    'conservative_daily': conservative_daily,
                    'daily_max': daily_max,
                    'volatility': volatility,
                    'trend_slope': slope,
                    'data_points': len(consumption_list),
                    'monitoring_frequency': monitoring_frequency
                },
                'scenarios': scenarios,
                'status': {
                    'level': status_level,
                    'urgency': urgency,
                    'color': status_color,
                    'message': message,
                    'action_required': status_level in ["CRITICAL", "URGENT", "WARNING"],
                    'days_to_reorder_threshold': days_to_reorder,
                    'days_to_minimum_threshold': days_to_minimum,
                    'recommended_order_date': (datetime.now() + timedelta(days=max(0, days_to_reorder - LEAD_TIME_DAYS))).strftime('%Y-%m-%d')
                },
                'recommendations': {
                    'immediate_actions': immediate_actions,
                    'order_recommendations': {
                        'optimal_quantity': optimal_quantity,
                        'minimum_quantity': minimum_quantity,
                        'recommended_order_date': (datetime.now() + timedelta(days=max(0, days_to_reorder - LEAD_TIME_DAYS))).strftime('%Y-%m-%d'),
                        'latest_order_date': (datetime.now() + timedelta(days=max(1, days_to_reorder - 1))).strftime('%Y-%m-%d')
                    }
                }
            }
        
        return jsonify({
            'timestamp': datetime.now().isoformat(),
            'forecast_results': forecast_results,
            'configuration': {
                'tank_capacities': TANK_CAPACITIES,
                'minimum_threshold_percent': MINIMUM_THRESHOLD * 100,
                'reorder_threshold_percent': REORDER_THRESHOLD * 100,
                'safety_buffer_days': SAFETY_BUFFER_DAYS,
                'lead_time_days': LEAD_TIME_DAYS,
                'analysis_lookback_days': ANALYSIS_LOOKBACK_DAYS
            }
        })
        
    except Exception as e:
        logger.error(f"Error in stock depletion forecast: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/business-intelligence/inventory-value', methods=['GET'])
def calculate_inventory_value():
    """Calculate current inventory value in rupees using latest rates from daily entries."""
    try:
        # Get current stock
        current_stock_response = get_current_stock()
        current_stock = current_stock_response.get_json()
        
        # Get latest rates from most recent daily entry (ordered by date, then by ID)
        latest_entry = DailyConsolidation.query.order_by(
            DailyConsolidation.date.desc(), 
            DailyConsolidation.id.desc()
        ).first()
        
        if latest_entry and latest_entry.ms_rate and latest_entry.hsd_rate and latest_entry.power_rate:
            # Use actual rates from the latest daily entry
            rates = {
                'ms_rate': float(latest_entry.ms_rate),
                'hsd_rate': float(latest_entry.hsd_rate),
                'power_rate': float(latest_entry.power_rate)
            }
            logger.info(f"Using rates from daily entry {latest_entry.date}: MS={rates['ms_rate']}, HSD={rates['hsd_rate']}, Power={rates['power_rate']}")
        else:
            # Fallback to default rates if no entry found or rates are missing
            rates = {
                'ms_rate': 106.50,
                'hsd_rate': 94.20,
                'power_rate': 89.30
            }
            logger.warning("No daily entry rates found, using default rates")
        
        # Calculate inventory values using the fetched rates
        inventory_value = {
            'MS': ((current_stock.get('ms1_tank', 0) + current_stock.get('ms2_tank', 0)) * rates['ms_rate']),
            'HSD': ((current_stock.get('hsd1_tank', 0) + current_stock.get('hsd2_tank', 0)) * rates['hsd_rate']),
            'Power': (current_stock.get('power1_tank', 0) * rates['power_rate'])
        }
        
        total_value = sum(inventory_value.values())
        
        logger.info(f"Calculated inventory values: MS=₹{inventory_value['MS']:,.2f}, HSD=₹{inventory_value['HSD']:,.2f}, Power=₹{inventory_value['Power']:,.2f}, Total=₹{total_value:,.2f}")
        
        return jsonify({
            'byFuel': inventory_value,
            'total': total_value,
            'rates': rates
        })
        
    except Exception as e:
        logger.error(f"Error calculating inventory value: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/business-intelligence/anomalies', methods=['GET'])
def detect_anomalies():
    """Detect statistical anomalies in sales and collections."""
    try:
        # Get last 30 days of data
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
        
        entries = DailyConsolidation.query.filter(
            DailyConsolidation.date >= start_date,
            DailyConsolidation.date <= end_date
        ).all()
        
        if len(entries) < 5:
            return jsonify([])  # Not enough data for anomaly detection
        
        # Calculate sales statistics
        sales_values = []
        for entry in entries:
            total_sales = (entry.ms_amount or 0) + (entry.hsd_amount or 0) + (entry.power_amount or 0)
            sales_values.append(total_sales)
        
        # Calculate mean and standard deviation
        mean_sales = sum(sales_values) / len(sales_values)
        variance = sum((x - mean_sales) ** 2 for x in sales_values) / len(sales_values)
        std_dev = variance ** 0.5
        
        anomalies = []
        
        for entry in entries:
            total_sales = (entry.ms_amount or 0) + (entry.hsd_amount or 0) + (entry.power_amount or 0)
            total_collections = (entry.cash_collections or 0) + (entry.card_collections or 0) + (entry.paytm_collections or 0)
            
            # Sales anomaly detection (2 standard deviations)
            if abs(total_sales - mean_sales) > 2 * std_dev:
                anomaly_type = 'High Sales' if total_sales > mean_sales else 'Low Sales'
                anomalies.append({
                    'date': entry.date.isoformat(),
                    'type': anomaly_type,
                    'description': f'Sales {"significantly above" if total_sales > mean_sales else "significantly below"} average (₹{total_sales:,.0f} vs ₹{mean_sales:,.0f})',
                    'severity': 'medium',
                    'value': total_sales
                })
            
            # Collection variance anomaly detection
            if total_sales > 0:
                variance_percent = ((total_collections - total_sales) / total_sales) * 100
                if abs(variance_percent) > 15:  # More than 15% variance
                    anomalies.append({
                        'date': entry.date.isoformat(),
                        'type': 'Collection Variance',
                        'description': f'Collection variance of {variance_percent:.1f}% detected',
                        'severity': 'critical' if abs(variance_percent) > 25 else 'medium',
                        'value': variance_percent
                    })
        
        # Sort by date (most recent first)
        anomalies.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify(anomalies[:10])  # Return top 10 most recent anomalies
        
    except Exception as e:
        logger.error(f"Error detecting anomalies: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)